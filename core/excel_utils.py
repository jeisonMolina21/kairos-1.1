"""
Utilidades para manejo de archivos Excel
Funciones comunes para formateo, estilos y manipulación de hojas
"""
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Optional
from core.config import styles as style_config


# =============================================================================
# CREACIÓN DE ESTILOS
# =============================================================================

def crear_fill(color: str, fill_type: str = "solid") -> PatternFill:
    """
    Crea un PatternFill con el color especificado
    
    Args:
        color: Código hexadecimal del color (sin #)
        fill_type: Tipo de fill (default: "solid")
        
    Returns:
        PatternFill configurado
    """
    return PatternFill(start_color=color, end_color=color, fill_type=fill_type)


def crear_fuente(
    nombre: str = None,
    tamaño: int = None,
    negrita: bool = False,
    color: str = "000000"
) -> Font:
    """
    Crea una fuente con los parámetros especificados
    
    Args:
        nombre: Nombre de la fuente (default: de configuración)
        tamaño: Tamaño de la fuente (default: de configuración)
        negrita: Si la fuente es negrita
        color: Color de la fuente en hex
        
    Returns:
        Font configurado
    """
    return Font(
        name=nombre or style_config.FUENTE_PRINCIPAL,
        size=tamaño or style_config.TAMAÑO_FUENTE_NORMAL,
        bold=negrita,
        color=color
    )


def crear_borde(
    lado: str = "all",
    estilo: str = "thin",
    color: str = "000000"
) -> Border:
    """
    Crea un borde con el estilo especificado
    
    Args:
        lado: Qué lados aplicar ("all", "box", "top", "bottom", "left", "right")
        estilo: Estilo del borde ("thin", "medium", "thick")
        color: Color del borde
        
    Returns:
        Border configurado
    """
    side = Side(style=estilo, color=color)
    
    if lado == "all" or lado == "box":
        return Border(left=side, right=side, top=side, bottom=side)
    elif lado == "top":
        return Border(top=side)
    elif lado == "bottom":
        return Border(bottom=side)
    elif lado == "left":
        return Border(left=side)
    elif lado == "right":
        return Border(right=side)
    else:
        return Border(left=side, right=side, top=side, bottom=side)


def crear_alineacion(
    horizontal: str = "center",
    vertical: str = "center",
    wrap_text: bool = False
) -> Alignment:
    """
    Crea una alineación con los parámetros especificados
    
    Args:
        horizontal: Alineación horizontal ("left", "center", "right")
        vertical: Alineación vertical ("top", "center", "bottom")
        wrap_text: Si el texto debe ajustarse
        
    Returns:
        Alignment configurado
    """
    return Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap_text
    )


# =============================================================================
# APLICACIÓN DE ESTILOS A CELDAS
# =============================================================================

def aplicar_estilo_celda(
    celda,
    fill: Optional[PatternFill] = None,
    font: Optional[Font] = None,
    border: Optional[Border] = None,
    alignment: Optional[Alignment] = None
):
    """
    Aplica estilos a una celda
    
    Args:
        celda: Celda de openpyxl
        fill: PatternFill a aplicar
        font: Font a aplicar
        border: Border a aplicar
        alignment: Alignment a aplicar
    """
    if fill:
        celda.fill = fill
    if font:
        celda.font = font
    if border:
        celda.border = border
    if alignment:
        celda.alignment = alignment


def aplicar_estilo_rango(
    ws: Worksheet,
    rango: str,
    fill: Optional[PatternFill] = None,
    font: Optional[Font] = None,
    border: Optional[Border] = None,
    alignment: Optional[Alignment] = None
):
    """
    Aplica estilos a un rango de celdas
    
    Args:
        ws: Worksheet de openpyxl
        rango: Rango en formato Excel (ej: "A1:C3")
        fill: PatternFill a aplicar
        font: Font a aplicar
        border: Border a aplicar
        alignment: Alignment a aplicar
    """
    for fila in ws[rango]:
        for celda in fila:
            aplicar_estilo_celda(celda, fill, font, border, alignment)


# =============================================================================
# CONFIGURACIÓN DE COLUMNAS
# =============================================================================

def ajustar_ancho_columna(ws: Worksheet, columna: int, ancho: float):
    """
    Ajusta el ancho de una columna
    
    Args:
        ws: Worksheet
        columna: Número de columna (1-indexed)
        ancho: Ancho en unidades de Excel
    """
    letter = get_column_letter(columna)
    ws.column_dimensions[letter].width = ancho


def ajustar_anchos_columnas(ws: Worksheet, anchos: List[float]):
    """
    Ajusta el ancho de múltiples columnas
    
    Args:
        ws: Worksheet
        anchos: Lista de anchos para cada columna
    """
    for idx, ancho in enumerate(anchos, 1):
        ajustar_ancho_columna(ws, idx, ancho)


def auto_ajustar_columnas(ws: Worksheet, min_width: float = 10, max_width: float = 50):
    """
    Auto-ajusta el ancho de todas las columnas basándose en el contenido
    
    Args:
        ws: Worksheet
        min_width: Ancho mínimo
        max_width: Ancho máximo
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


# =============================================================================
# CREACIÓN DE ENCABEZADOS
# =============================================================================

def crear_encabezado(
    ws: Worksheet,
    fila: int,
    encabezados: List[str],
    color_fondo: str = None,
    color_texto: str = "FFFFFF"
):
    """
    Crea una fila de encabezado con estilo
    
    Args:
        ws: Worksheet
        fila: Número de fila donde crear el encabezado
        encabezados: Lista de textos para el encabezado
        color_fondo: Color de fondo (hex), default del config
        color_texto: Color del texto (hex)
    """
    if color_fondo is None:
        color_fondo = style_config.AZUL_PRINCIPAL
    
    fill = crear_fill(color_fondo)
    font = crear_fuente(
        tamaño=style_config.TAMAÑO_FUENTE_ENCABEZADO,
        negrita=True,
        color=color_texto
    )
    border = crear_borde("all", "thin")
    alignment = crear_alineacion()
    
    for col_idx, texto in enumerate(encabezados, 1):
        celda = ws.cell(row=fila, column=col_idx, value=texto)
        aplicar_estilo_celda(celda, fill, font, border, alignment)


def crear_titulo(
    ws: Worksheet,
    fila: int,
    columna_inicio: int,
    columna_fin: int,
    titulo: str,
    color_fondo: str = None
):
    """
    Crea un título combinado en varias columnas
    
    Args:
        ws: Worksheet
        fila: Número de fila
        columna_inicio: Columna de inicio (1-indexed)
        columna_fin: Columna de fin (1-indexed)
        titulo: Texto del título
        color_fondo: Color de fondo
    """
    if color_fondo is None:
        color_fondo = style_config.AZUL_PRINCIPAL
    
    # Escribir el título en la primera celda
    celda = ws.cell(row=fila, column=columna_inicio, value=titulo)
    
    # Combinar celdas
    ws.merge_cells(
        start_row=fila,
        start_column=columna_inicio,
        end_row=fila,
        end_column=columna_fin
    )
    
    # Aplicar estilos
    fill = crear_fill(color_fondo)
    font = crear_fuente(
        tamaño=style_config.TAMAÑO_FUENTE_ENCABEZADO + 2,
        negrita=True,
        color="FFFFFF"
    )
    alignment = crear_alineacion()
    border = crear_borde("all", "medium")
    
    aplicar_estilo_celda(celda, fill, font, border, alignment)


# =============================================================================
# HELPERS DE FORMATO
# =============================================================================

def formatear_como_moneda(ws: Worksheet, columna: int, fila_inicio: int, fila_fin: int):
    """
    Formatea un rango de celdas como moneda
    
    Args:
        ws: Worksheet
        columna: Número de columna
        fila_inicio: Fila de inicio
        fila_fin: Fila de fin
    """
    for fila in range(fila_inicio, fila_fin + 1):
        celda = ws.cell(row=fila, column=columna)
        celda.number_format = '$#,##0.00'


def formatear_como_porcentaje(ws: Worksheet, columna: int, fila_inicio: int, fila_fin: int):
    """
    Formatea un rango de celdas como porcentaje
    
    Args:
        ws: Worksheet
        columna: Número de columna
        fila_inicio: Fila de inicio
        fila_fin: Fila de fin
    """
    for fila in range(fila_inicio, fila_fin + 1):
        celda = ws.cell(row=fila, column=columna)
        celda.number_format = '0.00%'


def formatear_como_tiempo(ws: Worksheet, columna: int, fila_inicio: int, fila_fin: int):
    """
    Formatea un rango de celdas como tiempo (HH:MM:SS)
    
    Args:
        ws: Worksheet
        columna: Número de columna
        fila_inicio: Fila de inicio
        fila_fin: Fila de fin
    """
    for fila in range(fila_inicio, fila_fin + 1):
        celda = ws.cell(row=fila, column=columna)
        celda.number_format = 'HH:MM:SS'


# =============================================================================
# HELPERS DE NAVEGACIÓN
# =============================================================================

def obtener_ultima_fila(ws: Worksheet) -> int:
    """
    Obtiene el número de la última fila con datos
    
    Args:
        ws: Worksheet
        
    Returns:
        Número de la última fila
    """
    return ws.max_row


def obtener_ultima_columna(ws: Worksheet) -> int:
    """
    Obtiene el número de la última columna con datos
    
    Args:
        ws: Worksheet
        
    Returns:
        Número de la última columna
    """
    return ws.max_column


# =============================================================================
# GESTIÓN DE HOJAS
# =============================================================================

def crear_o_limpiar_hoja(wb: Workbook, nombre_hoja: str) -> Worksheet:
    """
    Crea una hoja nueva o limpia una existente
    
    Args:
        wb: Workbook
        nombre_hoja: Nombre de la hoja
        
    Returns:
        Worksheet creada o limpiada
    """
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(nombre_hoja)
    
    return ws


def reordenar_hojas(wb: Workbook, orden: List[str]):
    """
    Reordena las hojas de un workbook según el orden especificado
    
    Args:
        wb: Workbook
        orden: Lista con nombres de hojas en el orden deseado
    """
    # Mover hojas al final en orden inverso
    for nombre_hoja in reversed(orden):
        if nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            wb.move_sheet(hoja, offset=len(wb.sheetnames) - 1)
    
    # Invertir el orden final
    wb._sheets.reverse()


# =============================================================================
# ESCRITURA DE DATOS
# =============================================================================

def escribir_fila_datos(
    ws: Worksheet,
    fila: int,
    datos: List,
    columna_inicio: int = 1,
    aplicar_borders: bool = True
):
    """
    Escribe una fila de datos
    
    Args:
        ws: Worksheet
        fila: Número de fila
        datos: Lista de valores a escribir
        columna_inicio: Columna de inicio (default 1)
        aplicar_borders: Si se aplican bordes
    """
    for col_idx, valor in enumerate(datos, columna_inicio):
        celda = ws.cell(row=fila, column=col_idx, value=valor)
        
        if aplicar_borders:
            celda.border = crear_borde("all", "thin")
        celda.alignment = crear_alineacion()
