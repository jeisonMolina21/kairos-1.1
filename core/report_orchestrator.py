"""
Orquestador de generación de reportes Excel
Coordina la creación de todas las hojas del reporte de asistencias
"""
import os
import openpyxl
from openpyxl import Workbook
import pandas as pd
from typing import Optional
from core.config import output, paths


class ReportOrchestrator:
    """Orquesta la generación de reportes Excel"""
    
    def __init__(self, temp_dir: str, output_dir: str = None):
        """
        Inicializa el orquestador
        
        Args:
            temp_dir: Directorio temporal con archivos JSON
            output_dir: Directorio de salida (default: de configuración)
        """
        self.temp_dir = temp_dir
        self.output_dir = output_dir or paths.OUTPUT_DIR
        self.reporte_path = os.path.join(self.output_dir, output.REPORTE_COMPLETO_FILENAME)
        
        # Rutas a archivos temporales
        self.resumen_path = os.path.join(temp_dir, output.RESUMEN_ASISTENCIAS_FILENAME)
        self.statistics_path = os.path.join(temp_dir, output.STATISTICS_FILENAME)
        self.datos_limpios_path = os.path.join(self.output_dir, output.DATOS_LIMPIOS_FILENAME)
    
    def generar_reporte_completo(self) -> str:
        """
        Genera el reporte Excel completo con todas las hojas
        
        Returns:
            Ruta al archivo Excel generado
        """
        print("\\n📊 Generando reporte Excel completo...")
        
        # Crear archivo base
        self._crear_archivo_base()
        
        # Generar cada hoja
        self._generar_hoja_diario()
        self._generar_hoja_resumen_semanal()
        self._generar_hoja_permisos()
        self._generar_hoja_horarios_programados()
        self._generar_hoja_explicacion()
        self._generar_hoja_registros_brutos()
        
        # Reordenar hojas
        self._reordenar_hojas()
        
        print(f"✅ Reporte completo generado: {self.reporte_path}")
        return self.reporte_path
    
    def _crear_archivo_base(self):
        """Crea el archivo Excel base"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Diario"
            wb.save(self.reporte_path)
            print("✅ Archivo base creado con hoja 'Diario'")
        except Exception as e:
            print(f"❌ Error al crear archivo base: {e}")
            raise
    
    def _generar_hoja_diario(self):
        """Genera la hoja de resumen diario"""
        try:
            from reports.excel_diario import agregar_hoja_diario
            
            agregar_hoja_diario(
                resumen_path=self.resumen_path,
                statistics_path=self.statistics_path,
                salida_excel=self.reporte_path
            )
            print("✅ Hoja 'Diario' generada")
        except Exception as e:
            print(f"❌ Error al generar hoja Diario: {e}")
            import traceback
            print(traceback.format_exc())
    
    def _generar_hoja_resumen_semanal(self):
        """Genera la hoja de resumen semanal"""
        try:
            from reports.excel_resumen import generar_excel_resumen_semanal
            
            generar_excel_resumen_semanal(
                resumen_path=self.resumen_path,
                statistics_path=self.statistics_path,
                salida_path=self.reporte_path
            )
            print("✅ Hoja 'Resumen Semanal' generada")
        except Exception as e:
            print(f"❌ Error al generar resumen semanal: {e}")
            import traceback
            print(traceback.format_exc())
    
    def _generar_hoja_permisos(self):
        """Genera la hoja de permisos"""
        try:
            from reports.excel_permisos import agregar_hoja_permisos
            
            agregar_hoja_permisos(
                statistics_path=self.statistics_path,
                salida_excel=self.reporte_path
            )
            print("✅ Hoja 'Permisos' generada")
        except Exception as e:
            print(f"❌ Error al generar hoja Permisos: {e}")
    
    def _generar_hoja_horarios_programados(self):
        """Genera la hoja de horarios programados"""
        try:
            from reports.excel_horarios_programados import agregar_hoja_horarios_programados
            
            agregar_hoja_horarios_programados(
                statistics_path=self.statistics_path,
                salida_excel=self.reporte_path
            )
            print("✅ Hoja 'Horarios Programados' generada")
        except Exception as e:
            print(f"❌ Error al generar hoja Horarios Programados: {e}")
    
    def _generar_hoja_explicacion(self):
        """Genera la hoja de explicación"""
        try:
            from reports.excel_explicacion import agregar_hoja_explicacion
            
            agregar_hoja_explicacion(salida_excel=self.reporte_path)
            print("✅ Hoja 'Explicación' generada")
        except Exception as e:
            print(f"❌ Error al generar hoja Explicación: {e}")
    
    def _generar_hoja_registros_brutos(self):
        """Genera la hoja de registros brutos"""
        try:
            df_brutos = pd.read_excel(self.datos_limpios_path)
            wb = openpyxl.load_workbook(self.reporte_path)

            # Eliminar hoja si ya existe
            if "Registros Brutos" in wb.sheetnames:
                del wb["Registros Brutos"]

            # Crear nueva hoja
            ws = wb.create_sheet("Registros Brutos")
            
            # Escribir encabezados
            for col_idx, col_name in enumerate(df_brutos.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Escribir datos
            for r_idx, row in enumerate(df_brutos.values, 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            wb.save(self.reporte_path)
            print("✅ Hoja 'Registros Brutos' generada")
        except Exception as e:
            print(f"❌ Error al generar hoja Registros Brutos: {e}")
    
    def _reordenar_hojas(self):
        """Reordena las hojas según el orden especificado en la configuración"""
        try:
            print("\\n🔀 Reordenando hojas...")
            wb = openpyxl.load_workbook(self.reporte_path)
            
            # Usar orden desde configuración
            orden_hojas = output.ORDEN_HOJAS_EXCEL
            
            # Reordenar hojas moviéndolas al final en orden inverso
            for nombre_hoja in reversed(orden_hojas):
                if nombre_hoja in wb.sheetnames:
                    hoja = wb[nombre_hoja]
                    wb.move_sheet(hoja, offset=len(wb.sheetnames) - 1)
            
            # Invertir el orden final
            wb._sheets.reverse()
            
            wb.save(self.reporte_path)
            print("✅ Hojas reordenadas correctamente:")
            for i, hoja in enumerate(wb.sheetnames, 1):
                print(f"   {i}. {hoja}")
        except Exception as e:
            print(f"⚠️ Error al reordenar hojas: {e}")
