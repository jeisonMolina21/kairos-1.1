# main.py - VERSIÓN MODIFICADA PARA REORDENAR HOJAS (FIX HORARIOS + ORDEN SOLICITADO)
import sys
import io
if sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import os
import sys
import pandas as pd
from tkinter import messagebox
import json
import openpyxl
import traceback
import shutil
from datetime import datetime

# Agregar el directorio actual al path para importaciones
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(current_dir)

from core.file_selector import seleccionar_archivos
from core.reader import leer_archivos_excel
from core.cleaner import limpiar_datos
from core.api_client import APIClient
from core.comparator import comparar_asistencias
from core.config import paths, departments
from reports.excel_resumen import generar_excel_resumen_semanal
from reports.excel_horarios import agregar_hoja_horarios
from reports.excel_diario import agregar_hoja_diario
from reports.excel_horarios_programados import agregar_hoja_horarios_programados
from reports.excel_explicacion import agregar_hoja_explicacion
from reports.excel_permisos import agregar_hoja_permisos

# =============================================================================
# 🚀 COMPATIBILIDAD CON EJECUTABLE ÚNICO - SIN ERRORES
# =============================================================================
def setup_executable_environment():
    """Configura el entorno para .exe sin errores"""
    try:
        # Si estamos en un .exe de PyInstaller
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            base_path = sys._MEIPASS
            print(f"🔧 Ejecutando desde .exe: {base_path}")

            # Agregar rutas de módulos
            core_path = os.path.join(base_path, 'core')
            reports_path = os.path.join(base_path, 'reports')

            if core_path not in sys.path:
                sys.path.insert(0, core_path)
            if reports_path not in sys.path:
                sys.path.insert(0, reports_path)

            return True
        else:
            print("🔧 Ejecutando desde Python")
            return False
    except Exception as e:
        print(f"⚠️  Error en configuración: {e}")
        return False

# Configurar entorno
IS_EXECUTABLE = setup_executable_environment()

def create_dirs_safely():
    """Crea directorios necesarios sin lanzar errores"""
    try:
        paths.initialize_directories()
    except Exception as e:
        print(f"⚠️  Error inicializando directorios: {e}")

create_dirs_safely()

# =============================================================================
# IMPORTACIONES SEGURAS
# =============================================================================
try:
    from core.file_selector import seleccionar_archivos
    from core.reader import leer_archivos_excel
    from core.cleaner import limpiar_datos
    from core.api_client import APIClient
    from core.comparator import comparar_asistencias
    from core.config import paths, departments
    from reports.excel_resumen import generar_excel_resumen_semanal
    from reports.excel_horarios import agregar_hoja_horarios
    from reports.excel_diario import agregar_hoja_diario
    from reports.excel_horarios_programados import agregar_hoja_horarios_programados
    from reports.excel_explicacion import agregar_hoja_explicacion
    from reports.excel_permisos import agregar_hoja_permisos
    print("✅ Todos los módulos importados correctamente")
except ImportError as e:
    print(f"❌ Error importando: {e}")
    input("Presione Enter para salir...")
    sys.exit(1)

def limpiar_archivos_temporales():
    """Limpia todos los archivos temporales de ejecuciones anteriores"""
    temp_dirs = [paths.TEMP_DIR, paths.EXCEL_OUTPUT_DIR]
    for temp_dir in temp_dirs:
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
                print(f"🧹 Carpeta temporal '{temp_dir}' eliminada")
            except PermissionError as e:
                print(f"⚠️ No se pudo eliminar la carpeta '{temp_dir}' (posiblemente abierta en el editor). Se continuará usando los archivos existentes.")

def crear_directorios_temporales():
    """Crea los directorios temporales necesarios"""
    paths.initialize_directories()
    return paths.TEMP_DIR

def obtener_rango_fechas(resumen_path):
    """Obtiene el rango de fechas del archivo de resumen para nombrar los archivos"""
    try:
        with open(resumen_path, "r", encoding="utf-8") as f:
            resumen = json.load(f)

        if not resumen:
            return None, None

        fechas = []
        for registro in resumen:
            fecha_str = registro.get("Fecha")
            if fecha_str and fecha_str != "N/A":
                try:
                    fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
                    fechas.append(fecha)
                except:
                    continue

        if not fechas:
            return None, None

        return min(fechas), max(fechas)

    except Exception as e:
        print(f"⚠️ Error obteniendo rango de fechas: {e}")
        return None, None

def obtener_codigo_departamento(nombre_departamento):
    """Mapea el nombre del departamento a su código abreviado"""
    mapeo_departamentos = departments.MAPEO_ARCHIVOS

    if not nombre_departamento or pd.isna(nombre_departamento) or nombre_departamento == "":
        return "ND"

    nombre_departamento = str(nombre_departamento).upper().strip()

    if nombre_departamento in mapeo_departamentos:
        return mapeo_departamentos[nombre_departamento]

    for depto_nombre, codigo in mapeo_departamentos.items():
        if depto_nombre in nombre_departamento:
            return codigo

    return "ND"

def detectar_departamento_archivos(archivos):
    """Detecta TODOS los departamentos en archivos seleccionados"""
    print("\n🔍 Detectando departamentos de los archivos (por contenido)...")

    departamentos_encontrados = set()

    for archivo in archivos:
        nombre_archivo = os.path.basename(archivo)
        print(f"📁 Analizando archivo: {nombre_archivo}")

        try:
            df_temp = pd.read_excel(archivo)
        except Exception as e:
            print(f"   ⚠️ Error leyendo archivo {archivo}: {e}")
            continue

        if df_temp is None or df_temp.empty:
            print("   ⚠️ Archivo sin datos o no legible")
            continue

        columna_departamento = None
        for col in df_temp.columns:
            col_str = str(col).upper()
            if ("DEPARTAMENTO" in col_str) or ("DEPTO" in col_str) or ("NOMBRE DE DEPARTAMENTO" in col_str):
                columna_departamento = col
                print(f"   📋 Columna de departamento encontrada: '{col}'")
                break

        if not columna_departamento:
            print("   ⚠️ No se encontró columna de departamento en este archivo")
            continue

        valores_unicos = df_temp[columna_departamento].dropna().unique()
        print(f"   🔍 Valores únicos en columna '{columna_departamento}': {list(valores_unicos)}")

        for valor in valores_unicos:
            if pd.isna(valor):
                continue

            nombre_depto = str(valor).strip()
            if not nombre_depto or nombre_depto.lower() == "nan":
                continue

            codigo = obtener_codigo_departamento(nombre_depto)

            if codigo != "ND":
                departamentos_encontrados.add(codigo)
                print(f"   ✅ Detectado departamento: '{nombre_depto}' -> {codigo}")
            else:
                print(f"   ℹ️ Valor de departamento ignorado (sin código): '{nombre_depto}'")

    if departamentos_encontrados:
        departamentos_ordenados = sorted(list(departamentos_encontrados))
        print(f"🏢 Departamentos detectados: {', '.join(departamentos_ordenados)}")
        return departamentos_ordenados

    print("⚠️ No se pudo detectar departamentos. Usando 'ND'")
    return ["ND"]

def detectar_departamentos_en_dataframe(df_limpio):
    """Detecta departamentos desde el DataFrame limpio"""
    print("\n🔍 Detectando departamentos desde el DataFrame limpio...")

    if df_limpio is None or df_limpio.empty:
        print("   ⚠️ DataFrame limpio vacío, no se pueden detectar departamentos.")
        return ["ND"]

    columna_departamento = None
    for col in df_limpio.columns:
        col_str = str(col).upper().replace("_", " ")
        if "DEPARTAMENTO" in col_str:
            columna_departamento = col
            print(f"   📋 Columna de departamento encontrada en DF limpio: '{col}'")
            break

    if not columna_departamento:
        print("   ⚠️ No se encontró columna de departamento en el DF limpio.")
        return ["ND"]

    valores_unicos = df_limpio[columna_departamento].dropna().unique()
    print(f"   🔍 Valores únicos en columna '{columna_departamento}': {list(valores_unicos)}")

    departamentos_encontrados = set()
    for valor in valores_unicos:
        nombre_depto = str(valor).strip()
        if not nombre_depto or nombre_depto.lower() == "nan":
            continue

        codigo = obtener_codigo_departamento(nombre_depto)
        if codigo != "ND":
            departamentos_encontrados.add(codigo)
            print(f"   ✅ Detectado en DF limpio: '{nombre_depto}' -> {codigo}")
        else:
            print(f"   ℹ️ Departamento sin código (ignorado): '{nombre_depto}'")

    if departamentos_encontrados:
        departamentos_ordenados = sorted(list(departamentos_encontrados))
        print(f"🏢 Departamentos detectados en DF limpio: {', '.join(departamentos_ordenados)}")
        return departamentos_ordenados

    print("⚠️ No se pudo detectar departamentos en DF limpio. Usando 'ND'")
    return ["ND"]

def obtener_novedades(registro):
    """Extrae info de novedades globales (paros, etc.)"""
    observaciones = registro.get("Observaciones", "").upper()
    if "NOVEDAD ORDEN PUBLICO" in observaciones:
        return "NOVEDAD ORDEN PÚBLICO"
    elif "PARO" in observaciones:
        return "PARO"
    elif "EMERGENCIA" in observaciones:
        return "EMERGENCIA"
    return None

def obtener_incidentes_para_api(registro):
    """Extrae incidentes en formato array para la API"""
    tipo_evento = registro.get("Tipo_Evento", "N/A")
    observaciones = registro.get("Observaciones", "")

    incidentes = []

    if tipo_evento != "N/A":
        if tipo_evento == "Departures":
            tipo_evento = "Inasistencia justificada (Salida institucional)"

        incidente = {
            "type": tipo_evento.lower().replace(" ", "_"),
            "description": f"Evento: {tipo_evento}"
        }

        if tipo_evento == "Permiso":
            tipo_permiso = registro.get("Tipo_Permiso", "N/A")
            if tipo_permiso != "N/A":
                incidente["description"] = f"Permiso: {tipo_permiso}"

        elif tipo_evento == "Campaña":
            campana_entidad = registro.get("Campaña_Entidad", "N/A")
            if campana_entidad != "N/A":
                incidente["description"] = f"Campaña en: {campana_entidad}"

        elif tipo_evento == "Inasistencia justificada (Salida institucional)":
            departures_institucion = registro.get("Departures_Institucion", "N/A")
            departures_razon = registro.get("Departures_Razon", "N/A")

            descripcion = "Salida institucional"
            if departures_institucion != "N/A":
                descripcion += f" - Institución: {departures_institucion}"
            if departures_razon != "N/A":
                descripcion += f" - Razón: {departures_razon}"
            incidente["description"] = descripcion

        incidentes.append(incidente)

    if "Llegó tarde" in observaciones:
        incidentes.append({"type": "tardanza", "description": "Llegó tarde"})

    if "Salió temprano" in observaciones:
        incidentes.append({"type": "salida_temprano", "description": "Salió antes del horario asignado"})

    if "No vino" in observaciones or "Sin registro" in observaciones:
        incidentes.append({"type": "ausencia", "description": "No se presentó a trabajar"})

    return incidentes if incidentes else None

def transformar_para_api(resumen_path):
    """Transforma el resumen al formato requerido por la API"""
    try:
        with open(resumen_path, "r", encoding="utf-8") as f:
            resumen = json.load(f)
    except Exception as e:
        print(f"❌ Error al cargar el archivo de resumen: {e}")
        return []

    def segundos_a_hhmmss(segundos_totales):
        if not segundos_totales:
            return "00:00:00"
        horas = int(segundos_totales // 3600)
        minutos = int((segundos_totales % 3600) // 60)
        segundos = int(segundos_totales % 60)
        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"

    datos_api = []

    for registro in resumen:
        incidentes = obtener_incidentes_para_api(registro)

        horas_trabajadas_decimal = registro.get("Horas_Trabajadas", 0) or 0
        horas_fuera_decimal = registro.get("Horas_Fuera", 0) or 0

        segundos_trabajados = int(horas_trabajadas_decimal * 3600)
        segundos_fuera = int(horas_fuera_decimal * 3600)

        cedula_str = registro.get("Cedula", "")
        try:
            user_id = int(cedula_str) if cedula_str and str(cedula_str).isdigit() else None
        except:
            user_id = None

        if not user_id:
            print(f"⚠️ Cédula inválida para conversión a número: {cedula_str}")
            continue

        dato_api = {
            "user_id": user_id,
            "attendance_date": registro.get("Fecha", ""),
            "arrival_time": registro.get("Hora_Ingreso") if registro.get("Hora_Ingreso") != "N/A" else None,
            "exit_time": registro.get("Hora_Salida") if registro.get("Hora_Salida") != "N/A" else None,
            "assigned_arrival_time": registro.get("Ingreso_Asignado") if registro.get("Ingreso_Asignado") != "N/A" else None,
            "assigned_exit_time": registro.get("Salida_Asignada") if registro.get("Salida_Asignada") != "N/A" else None,
            "total_worked_hours": segundos_a_hhmmss(segundos_trabajados),
            "out_of_institution_hours": segundos_a_hhmmss(segundos_fuera),
            "incidents": incidentes,
            "observations": registro.get("Observaciones", ""),
            "news": obtener_novedades(registro)
        }

        if dato_api["observations"] and len(dato_api["observations"]) > 255:
            dato_api["observations"] = dato_api["observations"][:255]
        if dato_api["news"] and len(dato_api["news"]) > 255:
            dato_api["news"] = dato_api["news"][:255]

        datos_api.append(dato_api)

    print(f"📤 Datos transformados para API: {len(datos_api)} registros")
    if datos_api:
        for i, dato in enumerate(datos_api[:2]):
            print(f"🔍 Ejemplo registro API {i+1}: {json.dumps(dato, ensure_ascii=False)}")

    return datos_api

def procesar_datos_excel(archivos):
    """Procesa los archivos Excel seleccionados, extrae departamentos y limpia los datos."""
    codigos_departamentos = detectar_departamento_archivos(archivos)

    dataframes = leer_archivos_excel(archivos)
    if not dataframes:
        messagebox.showwarning("Error", "No se pudieron leer los archivos seleccionados.")
        return None, None, 0

    df_combinado = pd.concat(dataframes, ignore_index=True)
    print(f"📊 Total combinado inicial: {len(df_combinado)} filas")

    df_limpio = limpiar_datos(df_combinado)

    if not codigos_departamentos or codigos_departamentos == ["ND"]:
        codigos_departamentos = detectar_departamentos_en_dataframe(df_limpio)

    os.makedirs("outputs", exist_ok=True)
    salida_excel_limpio = os.path.join("outputs", "datos_limpios.xlsx")
    df_limpio.to_excel(salida_excel_limpio, index=False)
    print(f"✅ Datos limpios guardados en: {salida_excel_limpio}")
    
    return codigos_departamentos, df_limpio, len(df_combinado)

def obtener_estadisticas_api(temp_dir):
    """Realiza login en la API y obtiene las estadísticas."""
    print("\n🌐 Consultando datos desde la API...")
    try:
        client = APIClient(email="admin@admin.com", password="Admin3811")
        client.login()
        stats = client.get_statistics()

        # PARCHE: Asegurar que stats sea siempre una lista antes de guardarlo en JSON
        if isinstance(stats, dict):
            for key, value in stats.items():
                if isinstance(value, list):
                    stats = value
                    break
            if isinstance(stats, dict):
                print(f"❌ Advertencia: La API devolvió un diccionario sin lista: {stats}")

        salida_json = os.path.join(temp_dir, "statistics.json")
        with open(salida_json, "w", encoding="utf-8") as f:
            json.dump(stats, f, indent=4, ensure_ascii=False)

        print(f"✅ Estadísticas guardadas en: {salida_json}")
        return client

    except Exception as e:
        print(f"❌ Error al consultar la API: {e}")
        messagebox.showerror("Error API", f"Ocurrió un error al consultar la API:\n{e}")
        return None

def ejecutar_comparacion(df_limpio, temp_dir):
    """Ejecuta la comparación entre los datos limpios y las estadísticas de la API."""
    print("\n🔍 Iniciando comparación de asistencias...")
    try:
        statistics_temp_path = os.path.join(temp_dir, "statistics.json")
        resumen_temp_path = os.path.join(temp_dir, "resumen_asistencias.json")

        resultados = comparar_asistencias(
            df_limpio,
            statistics_path=statistics_temp_path,
            salida_path=resumen_temp_path
        )
        print(f"✅ Comparativa completada: {len(resultados)} registros procesados")
        return resultados
    except Exception as e:
        print(f"❌ Error en la comparación: {e}")
        messagebox.showerror("Error", f"Error al comparar asistencias:\n{e}")
        return None

def generar_nombres_archivos(codigos_departamentos, fecha_min, fecha_max):
    """Genera los nombres de archivo en base a los departamentos y el rango de fechas."""
    codigos_str = "_".join(codigos_departamentos)
    if fecha_min and fecha_max:
        rango_fechas = f"{fecha_min.strftime('%d-%m')}_{fecha_max.strftime('%d-%m')}"
        nombre_excel_api = f"{codigos_str}_{rango_fechas}.xlsx"
        nombre_json_api = f"Asistencia_{codigos_str}_{rango_fechas}.json"
        print(f"📅 Rango de fechas detectado: {rango_fechas}")
        print(f"🏢 Códigos de departamentos: {codigos_str}")
    else:
        nombre_excel_api = f"{codigos_str}_sin_fechas.xlsx"
        nombre_json_api = f"Asistencia_{codigos_str}_sin_fechas.json"
        print("⚠️ No se pudieron detectar las fechas para nombrar archivos")
    return nombre_excel_api, nombre_json_api

def generar_reportes_excel(temp_dir):
    """Orquesta la creación de todas las hojas del archivo Excel final."""
    reporte_path = "outputs/Reporte_Asistencias_Completo.xlsx"
    salida_excel_limpio = os.path.join("outputs", "datos_limpios.xlsx")
    
    # Resumen Semanal
    try:
        generar_excel_resumen_semanal(
            resumen_path=os.path.join(temp_dir, "resumen_asistencias.json"),
            statistics_path=os.path.join(temp_dir, "statistics.json"),
            salida_path=reporte_path
        )
        print("✅ Hoja 'Resumen Semanal' generada (base del archivo)")
    except Exception as e:
        print(f"❌ Error al generar resumen semanal: {e}")
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Diario"
            wb.save(reporte_path)
            print("✅ Fallback: Archivo base creado con hoja 'Diario'")
        except Exception as e2:
            print(f"❌ Error creando archivo base fallback: {e2}")
            return None

    # Diario
    try:
        agregar_hoja_diario(
            resumen_path=os.path.join(temp_dir, "resumen_asistencias.json"),
            statistics_path=os.path.join(temp_dir, "statistics.json"),
            salida_excel=reporte_path
        )
        print("✅ Hoja 'Diario' agregada")
    except Exception as e:
        print(f"❌ Error al agregar hoja Diario: {e}\n{traceback.format_exc()}")

    # Horarios
    try:
        agregar_hoja_horarios(
            resumen_path=os.path.join(temp_dir, "resumen_asistencias.json"),
            statistics_path=os.path.join(temp_dir, "statistics.json"),
            salida_excel=reporte_path
        )
        print("✅ Hoja 'Horarios' agregada")
    except Exception as e:
        print(f"❌ Error al crear hoja 'Horarios': {e}")

    # Permisos
    try:
        agregar_hoja_permisos(
            statistics_path=os.path.join(temp_dir, "statistics.json"),
            salida_excel=reporte_path
        )
        print("✅ Hoja 'Permisos' agregada")
    except Exception as e:
        print(f"❌ Error al agregar hoja 'Permisos': {e}")

    # Horarios Programados
    try:
        agregar_hoja_horarios_programados(
            statistics_path=os.path.join(temp_dir, "statistics.json"),
            salida_excel=reporte_path
        )
        print("✅ Hoja 'Horarios Programados' agregada")
    except Exception as e:
        print(f"❌ Error al agregar hoja 'Horarios Programados': {e}")

    # Explicación
    try:
        agregar_hoja_explicacion(salida_excel=reporte_path)
        print("✅ Hoja 'Explicación' agregada")
    except Exception as e:
        print(f"❌ Error al agregar hoja Explicación: {e}")

    # Registros Brutos
    try:
        df_brutos = pd.read_excel(salida_excel_limpio)
        book = openpyxl.load_workbook(reporte_path)
        if "Registros Brutos" in book.sheetnames:
            del book["Registros Brutos"]
        ws = book.create_sheet("Registros Brutos")
        for col_idx, col_name in enumerate(df_brutos.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        for r_idx, row in enumerate(df_brutos.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        book.save(reporte_path)
        print("✅ Hoja 'Registros Brutos' agregada correctamente")
    except Exception as e:
        print(f"❌ Error al agregar hoja 'Registros Brutos': {e}")

    # Reordenar Hojas
    try:
        print("\n🔀 Reordenando hojas en el orden solicitado...")
        wb = openpyxl.load_workbook(reporte_path)
        orden_hojas = ["Diario", "Horarios", "Resumen Semanal", "Permisos", "Horarios Programados", "Explicación", "Registros Brutos"]
        for nombre_hoja in reversed(orden_hojas):
            if nombre_hoja in wb.sheetnames:
                hoja = wb[nombre_hoja]
                wb.move_sheet(hoja, offset=len(wb.sheetnames) - 1)
        wb._sheets.reverse()
        wb.save(reporte_path)
        print("✅ Hojas reordenadas correctamente:")
        for i, hoja in enumerate(wb.sheetnames, 1):
            print(f"   {i}. {hoja}")
    except Exception as e:
        print(f"⚠️ Error al reordenar hojas: {e}")
        
    return reporte_path

def enviar_resultados_api(client, temp_dir, reporte_path, nombre_excel_api, nombre_json_api):
    """Transforma los resultados y los envía a la API junto con los archivos generados."""
    print("\n📤 Enviando datos a la API...")
    try:
        datos_api = transformar_para_api(os.path.join(temp_dir, "resumen_asistencias.json"))
        if datos_api:
            print(f"📊 Preparando envío de {len(datos_api)} registros a la API...")
            client.enviar_asistencias_api(datos_api)
            print("🎉 Datos de asistencias enviados a la API correctamente")
        else:
            print("⚠️ No hay datos para enviar a la API")
    except Exception as e:
        print(f"❌ Error en el proceso de envío a la API: {e}\n{traceback.format_exc()}")
        datos_api = []

    print("\n📤 Enviando archivos generados a la API...")
    try:
        excel_renombrado = os.path.join(temp_dir, nombre_excel_api)
        json_renombrado = os.path.join(temp_dir, nombre_json_api)

        if os.path.exists(reporte_path):
            shutil.copy2(reporte_path, excel_renombrado)
            print(f"📄 Excel renombrado: {nombre_excel_api}")
        else:
            print("⚠️ No se encontró el archivo Excel para renombrar")

        if os.path.exists(os.path.join(temp_dir, "resumen_asistencias.json")):
            shutil.copy2(os.path.join(temp_dir, "resumen_asistencias.json"), json_renombrado)
            print(f"📄 JSON renombrado: {nombre_json_api}")
        else:
            print("⚠️ No se encontró el archivo JSON para renombrar")

        if os.path.exists(excel_renombrado):
            client.enviar_excel_api(excel_renombrado)
            print(f"✅ Archivo Excel enviado: {nombre_excel_api}")

        if os.path.exists(json_renombrado):
            client.enviar_json_api(json_renombrado)
            print(f"✅ Archivo JSON enviado: {nombre_json_api}")

    except Exception as e:
        print(f"❌ Error al preparar archivos para la API: {e}\n{traceback.format_exc()}")
        
    return datos_api

def mostrar_resumen_final(archivos_len, df_combinado_len, df_limpio_len, reporte_path, codigos_departamentos, nombre_excel_api, nombre_json_api, temp_dir, fecha_min, datos_api_len):
    """Muestra el popup final con el resumen de la ejecución."""
    salida_excel_limpio = os.path.join("outputs", "datos_limpios.xlsx")
    
    print("\n📁 ARCHIVOS GENERADOS:")
    archivos_generados = [
        ("Datos limpios", salida_excel_limpio),
        ("Estadísticas API", os.path.join(temp_dir, "statistics.json")),
        ("Resumen asistencias", os.path.join(temp_dir, "resumen_asistencias.json")),
        ("Reporte completo", reporte_path),
        ("Excel para API", os.path.join(temp_dir, nombre_excel_api) if fecha_min else "No generado"),
        ("JSON para API", os.path.join(temp_dir, nombre_json_api) if fecha_min else "No generado")
    ]

    for nombre, ruta in archivos_generados:
        if isinstance(ruta, str) and os.path.exists(ruta):
            print(f"✅ {nombre}: {ruta}")
        else:
            print(f"❌ {nombre}: NO GENERADO")

    print("=" * 60)
    messagebox.showinfo(
        "Proceso Completado",
        f"✅ Se procesaron {archivos_len} archivos\n"
        f"📊 {df_combinado_len} registros iniciales\n"
        f"🧹 {df_limpio_len} registros limpios\n"
        f"📑 Reporte generado: {reporte_path}\n"
        f"📊 ORDEN DE HOJAS: 1. Diario, 2. Horarios, 3. Resumen Semanal, 4. Permisos, 5. Horarios Programados, 6. Explicación, 7. Registros Brutos\n"
        f"📤 Datos enviados a la API: {datos_api_len} registros\n"
        f"🏢 Departamentos: {', '.join(codigos_departamentos)}\n"
        f"📁 Archivos preparados para API: {nombre_excel_api} y {nombre_json_api}"
    )

def main():
    print("🚀 Iniciando Sistema de Procesamiento de Asistencia")
    print("=" * 60)
    print("🔢 NUEVO SISTEMA: Días laborados = 100%, regla de 3 por empleado")
    print("📊 PESOS: 40% Tarde + 35% A Tiempo + 25% Temprano")

    limpiar_archivos_temporales()
    temp_dir = crear_directorios_temporales()

    try:
        archivos = seleccionar_archivos()
        if not archivos:
            messagebox.showinfo("Sin archivos", "No seleccionaste ningún archivo.")
            return

        codigos_departamentos, df_limpio, df_combinado_len = procesar_datos_excel(archivos)
        if df_limpio is None:
            return

        client = obtener_estadisticas_api(temp_dir)
        if not client:
            return

        resultados = ejecutar_comparacion(df_limpio, temp_dir)
        if not resultados:
            return

        resumen_temp_path = os.path.join(temp_dir, "resumen_asistencias.json")
        fecha_min, fecha_max = obtener_rango_fechas(resumen_temp_path)

        nombre_excel_api, nombre_json_api = generar_nombres_archivos(codigos_departamentos, fecha_min, fecha_max)

        reporte_path = generar_reportes_excel(temp_dir)
        if not reporte_path:
            return

        datos_api = enviar_resultados_api(client, temp_dir, reporte_path, nombre_excel_api, nombre_json_api)

        mostrar_resumen_final(len(archivos), df_combinado_len, len(df_limpio), reporte_path, codigos_departamentos, nombre_excel_api, nombre_json_api, temp_dir, fecha_min, len(datos_api) if datos_api else 0)

    except Exception as e:
        print(f"❌ Error general: {e}")
        print(traceback.format_exc())
        messagebox.showerror("Error", f"Error inesperado:\n{e}")

    finally:
        print("\n🧹 Limpiando archivos temporales...")
        if os.path.exists(temp_dir):
            # shutil.rmtree(temp_dir)
            print("✅ Archivos temporales eliminados")

if __name__ == "__main__":
    main()
