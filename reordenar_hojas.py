"""
reordenar_hojas.py
Módulo para reordenar las hojas del Excel en el orden especificado
"""
from openpyxl import load_workbook

def reordenar_hojas_excel(archivo_path):
    """Reordena las hojas del Excel en el orden especificado"""
    print(f"\n🔄 Reordenando hojas en {archivo_path}...")
    
    try:
        wb = load_workbook(archivo_path)
        
        # Orden deseado (exactamente como lo pides)
        orden_deseado = [
            'Diario',
            'Resumen Semanal', 
            'Permisos',
            'Horarios Programados',
            'Explicación',
            'Registros Brutos'
        ]
        
        # Verificar qué hojas existen realmente
        hojas_existentes = wb.sheetnames
        print(f"📑 Hojas existentes: {hojas_existentes}")
        
        # Crear lista de hojas en el orden deseado (solo las que existen)
        hojas_ordenadas = []
        
        for nombre in orden_deseado:
            if nombre in hojas_existentes:
                hojas_ordenadas.append(wb[nombre])
        
        # Limpiar todas las hojas existentes
        for sheetname in hojas_existentes:
            if sheetname not in orden_deseado:
                del wb[sheetname]
        
        # Reorganizar workbook
        wb._sheets = hojas_ordenadas
        
        # Guardar cambios
        wb.save(archivo_path)
        
        # Verificar el orden final
        wb_final = load_workbook(archivo_path, read_only=True)
        print(f"✅ Hojas reordenadas: {', '.join(wb_final.sheetnames)}")
        
    except Exception as e:
        print(f"⚠️  Error reordenando hojas: {e}")
        print("ℹ️  Continuando sin reordenar...")

if __name__ == "__main__":
    # Ejemplo de uso
    reordenar_hojas_excel("outputs/Reporte_Asistencias_Completo.xlsx")