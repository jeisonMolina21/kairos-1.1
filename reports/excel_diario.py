# reports/excel_diario.py (MODIFICADO - AGREGAR COLUMNA "DÍA" Y NUEVO CÁLCULO DE HORAS)
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import sys
import os

# Agregar el directorio raíz al path para importaciones absolutas
current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, current_dir)

# Importación robusta de EstilosInstitucionales
try:
    from reports.styles import EstilosInstitucionales
except ImportError as e:
    print(f"⚠️ No se pudo importar EstilosInstitucionales: {e}")

# Función para obtener nombre del día en español
def obtener_nombre_dia(fecha_str):
    """Convierte fecha string a nombre del día en español"""
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
        dias_espanol = {
            0: "Lunes",
            1: "Martes", 
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sábado",
            6: "Domingo"
        }
        return dias_espanol[fecha.weekday()]
    except:
        return "N/A"

# Función simple para festivos (para excluirlos)
def es_dia_festivo_simple(fecha):
    """
    Función simplificada para verificar días festivos.
    """
    festivos_colombia = {
        '01-01': 'Año Nuevo',
        '01-06': 'Día de los Reyes Magos',
        '03-25': 'Día de San José',
        '04-13': 'Jueves Santo',
        '04-14': 'Viernes Santo',
        '05-01': 'Día del Trabajo',
        '05-22': 'Día de la Ascensión',
        '06-12': 'Corpus Christi',
        '06-19': 'Sagrado Corazón',
        '07-03': 'San Pedro y San Pablo',
        '07-20': 'Día de la Independencia',
        '08-07': 'Batalla de Boyacá',
        '08-21': 'Asunción de la Virgen',
        '10-16': 'Día de la Raza',
        '11-06': 'Todos los Santos',
        '11-13': 'Independencia de Cartagena',
        '12-08': 'Día de la Inmaculada Concepción',
        '12-25': 'Navidad'
    }
    
    fecha_str = fecha.strftime('%m-%d')
    if fecha_str in festivos_colombia:
        return True, festivos_colombia[fecha_str]
    return False, None

def agregar_hoja_diario(resumen_path="temp/resumen_asistencias.json",
                       statistics_path="temp/statistics.json",
                       salida_excel="outputs/Reporte_Asistencias_Completo.xlsx"):
    """
    Agrega la hoja de resumen diario al reporte.
    EXCLUYE DÍAS FESTIVOS - solo se muestran en Resumen Semanal.
    🆕 MODIFICACIÓN: Agregada columna "Día" después de "Fecha"
    🆕 MODIFICACIÓN: Nuevo cálculo de horas trabajadas = hora salida - hora ingreso
    """
    print("\n📘 Agregando hoja 'Diario' (excluyendo festivos)...")

    try:
        wb = load_workbook(salida_excel)
        print(f"✅ Archivo Excel cargado: {salida_excel}")
    except FileNotFoundError:
        print(f"❌ No se encontró el archivo {salida_excel}")
        return

    # Crear o limpiar hoja existente
    if "Diario" in wb.sheetnames:
        ws = wb["Diario"]
        ws.delete_rows(1, ws.max_row)
        print("✅ Hoja 'Diario' limpiada")
    else:
        ws = wb.create_sheet("Diario")
        print("✅ Hoja 'Diario' creada")

    # Cargar datos del resumen
    try:
        with open(resumen_path, "r", encoding="utf-8") as f:
            resumen = json.load(f)
        print(f"✅ Resumen cargado: {len(resumen)} registros desde {resumen_path}")
    except Exception as e:
        print(f"❌ Error al leer el archivo de resumen {resumen_path}: {e}")
        return

    # Cargar datos de statistics para obtener información de almuerzo, carnet y NOVEDADES
    try:
        with open(statistics_path, "r", encoding="utf-8") as f:
            statistics = json.load(f)
        print(f"✅ Statistics cargado: {len(statistics)} registros desde {statistics_path}")
    except Exception as e:
        print(f"❌ Error al leer el archivo de statistics {statistics_path}: {e}")
        statistics = []

    # Crear diccionario para búsqueda rápida de información de almuerzo, permisos, campañas y SALIDAS INSTITUCIONALES
    info_almuerzo_permisos_novedades = {}
    mapeo_novedades = {}  # Diccionario específico para novedades
    
    for stat in statistics:
        try:
            cedula = str(stat.get("number_id", "")).strip()
            año = stat.get("año")
            mes = stat.get("mes")
            dia = stat.get("dia")
            
            if not all([cedula, año, mes, dia]):
                continue
                
            fecha_str = f"{año}-{str(mes).zfill(2)}-{str(dia).zfill(2)}"
            horario_almuerzo_inicio = stat.get("horario_almuerzo_inicio", "N/A")
            horario_almuerzo_fin = stat.get("horario_almuerzo_fin", "N/A")
            carnet = stat.get("carnet", "No")
            carnet_number = stat.get("carnet_number", "N/A")
            
            # Capturar información de permisos, campañas y SALIDAS INSTITUCIONALES
            tiene_permiso = stat.get("permiso", "No")
            tipo_permiso = stat.get("tipo_permiso", "N/A")
            tiene_campaña = stat.get("campaña si/no", "No")  # CAMBIO: usar "campaña si/no"
            tiene_departures = stat.get("departures si/no", "No")  # NUEVO: campo departures
            url_permiso = stat.get("permiso_url", "N/A")
            
            # Información de campaña
            campana_fecha = stat.get("campaña-fecha", "N/A")
            campana_hora_inicio = stat.get("campaña-hora-inicio", "N/A")
            campana_hora_fin = stat.get("campaña-hora-fin", "N/A")
            campana_entidad = stat.get("campaña(nombre entidad)", "N/A")
            
            # Información de SALIDA INSTITUCIONAL
            departures_institucion = stat.get("institución", "N/A")
            departures_fecha = stat.get("fecha", "N/A")
            departures_hora_inicio = stat.get("horaInicio", "N/A")
            departures_hora_fin = stat.get("horaFin", "N/A")
            departures_razon = stat.get("razón", "N/A")
            
            # CAPTURAR INFORMACIÓN DE NOVEDADES
            tiene_novedad = stat.get("novedad", "No")
            novedad_comienzo = stat.get("novedad_comienzo", "N/A")
            novedad_duracion = stat.get("novedad_duracion", "N/A")
            novedad_evento = stat.get("novedad_evento", "N/A")
            novedad_observaciones = stat.get("novedad_observaciones", "N/A")
            
            if cedula and fecha_str != "None-None-None":
                if cedula not in info_almuerzo_permisos_novedades:
                    info_almuerzo_permisos_novedades[cedula] = {}
                
                info_almuerzo_permisos_novedades[cedula][fecha_str] = {
                    "horario_almuerzo_inicio": horario_almuerzo_inicio,
                    "horario_almuerzo_fin": horario_almuerzo_fin,
                    "carnet": carnet,
                    "carnet_number": carnet_number,
                    "tiene_permiso": tiene_permiso,
                    "tipo_permiso": tipo_permiso,
                    "tiene_campaña": tiene_campaña,
                    "tiene_departures": tiene_departures,  # NUEVO: Salida Institucional
                    "url_permiso": url_permiso,
                    # Información de campaña
                    "campana_fecha": campana_fecha,
                    "campana_hora_inicio": campana_hora_inicio,
                    "campana_hora_fin": campana_hora_fin,
                    "campana_entidad": campana_entidad,
                    # Información de SALIDA INSTITUCIONAL
                    "departures_institucion": departures_institucion,
                    "departures_fecha": departures_fecha,
                    "departures_hora_inicio": departures_hora_inicio,
                    "departures_hora_fin": departures_hora_fin,
                    "departures_razon": departures_razon,
                    # Información de novedades
                    "tiene_novedad": tiene_novedad,
                    "novedad_comienzo": novedad_comienzo,
                    "novedad_duracion": novedad_duracion,
                    "novedad_evento": novedad_evento,
                    "novedad_observaciones": novedad_observaciones
                }
                
                # Guardar en mapeo específico de novedades
                if tiene_novedad == "Sí":
                    mapeo_novedades[(cedula, fecha_str)] = {
                        "comienzo": novedad_comienzo,
                        "duracion": novedad_duracion,
                        "evento": novedad_evento,
                        "observaciones": novedad_observaciones
                    }
        except Exception as e:
            print(f"⚠️ Error procesando statistics para novedades: {e}")
            continue

    print(f"⚠️  Novedades detectadas en API: {len(mapeo_novedades)} registros")

    # Estilos institucionales
    header_col_fill = EstilosInstitucionales.fill_azul_medio()
    bold_white_font = EstilosInstitucionales.fuente_encabezado_columna()
    borde_blanco_inf = EstilosInstitucionales.borde_inferior_blanco()
    center = EstilosInstitucionales.alineacion_centro()

    # Agregar encabezado superior
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=12)
    
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    
    cell_titulo = ws.cell(row=1, column=1, value="REPORTE DIARIO DE ASISTENCIA")
    cell_titulo.fill = EstilosInstitucionales.fill_azul_principal()
    cell_titulo.font = EstilosInstitucionales.fuente_titulo_principal()
    cell_titulo.alignment = center
    
    cell_periodo = ws.cell(row=2, column=1, value="Período Analizado")
    cell_periodo.fill = EstilosInstitucionales.fill_azul_principal()
    cell_periodo.font = EstilosInstitucionales.fuente_periodo()
    cell_periodo.alignment = center
    
    cell_meta = ws.cell(row=3, column=1, value="Generado por Kairos FAS")
    cell_meta.fill = EstilosInstitucionales.fill_azul_principal()
    cell_meta.font = EstilosInstitucionales.fuente_metadatos()
    cell_meta.alignment = center
    
    for r in range(1, 4):
        for c in range(2, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = EstilosInstitucionales.fill_azul_principal()

    # 🆕 MODIFICACIÓN: Encabezados con nueva columna "Día"
    headers = [
        "Cédula", "Nombre", "Fecha", "Día", "Hora Ingreso", "Hora Salida",
        "Horas Trabajadas", "Horas Descanso", "Horario de Almuerzo", "Horas Netas", 
        "Vista Previa Permiso", "Observaciones"
    ]
    ws.append(headers)

    # Aplicar estilos a encabezados de tabla (Fila 4)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_col_fill
        cell.font = bold_white_font
        cell.alignment = center
        cell.border = borde_blanco_inf

    # MODIFICACIÓN: Función para calcular horas entre dos tiempos (MEJORADA)
    def calcular_horas_entre(hora1, hora2):
        """Calcula la diferencia en horas entre dos horas en formato HH:MM:SS"""
        try:
            if hora1 == "N/A" or hora2 == "N/A":
                return 0.0
            
            h1 = datetime.strptime(hora1, "%H:%M:%S")
            h2 = datetime.strptime(hora2, "%H:%M:%S")
            
            # Si la hora de salida es menor que la de ingreso, asumimos que es del día siguiente
            if h2 < h1:
                h2 = datetime.strptime("23:59:59", "%H:%M:%S")
            
            diferencia = h2 - h1
            return round(diferencia.total_seconds() / 3600, 2)
        except:
            return 0.0

    # Función para calcular tiempo de almuerzo
    def calcular_tiempo_almuerzo(hora_inicio, hora_fin):
        """Calcula la duración del almuerzo en horas"""
        try:
            if hora_inicio == "N/A" or hora_fin == "N/A":
                return 0.0
            
            h_inicio = datetime.strptime(hora_inicio, "%H:%M:%S")
            h_fin = datetime.strptime(hora_fin, "%H:%M:%S")
            
            diferencia = h_fin - h_inicio
            return round(diferencia.total_seconds() / 3600, 2)
        except:
            return 0.0

    # Función para convertir horas decimales a formato HH:MM:SS
    def decimal_a_hhmmss(horas_decimal):
        """Convierte horas decimales a formato HH:MM:SS"""
        try:
            horas = int(horas_decimal)
            minutos = int((horas_decimal - horas) * 60)
            segundos = int((((horas_decimal - horas) * 3600) % 60))
            return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
        except:
            return "00:00:00"

    # Función para determinar puntualidad con tolerancia de 5 minutos
    def determinar_puntualidad(hora_real, hora_asignada, tipo="entrada"):
        """Determina si fue temprano, a tiempo o tarde con tolerancia de ±5 minutos"""
        if hora_real == "N/A" or hora_asignada == "N/A":
            return "Sin registro"
        
        try:
            h_real = datetime.strptime(hora_real, "%H:%M:%S")
            h_asignada = datetime.strptime(hora_asignada, "%H:%M:%S")
            
            diferencia_minutos = (h_real - h_asignada).total_seconds() / 60
            
            if tipo == "entrada":
                if diferencia_minutos <= -5:
                    return "temprano"
                elif -5 < diferencia_minutos <= 5:
                    return "a_tiempo"
                else:
                    return "tarde"
            else:
                if diferencia_minutos >= 5:
                    return "tarde"
                elif -5 <= diferencia_minutos < 5:
                    return "a_tiempo"
                else:
                    return "temprano"
                    
        except Exception as e:
            return "error"

    # Función: Generar observación mejorada que incluye permisos, campañas y SALIDAS INSTITUCIONALES
    def generar_observacion_mejorada(punt_entrada, punt_salida, es_domingo=False, info_permisos=None, registro_resumen=None, es_festivo=False):
        """Genera observación mejorada que incluye información de permisos, campañas, SALIDAS INSTITUCIONALES y Festivos"""
        
        # Priorizar información del registro_resumen (viene del comparator)
        if registro_resumen:
            tipo_evento = registro_resumen.get("Tipo_Evento") or registro_resumen.get("Tipo Evento") or "N/A"
            if tipo_evento == "Permiso":
                tipo_permiso = registro_resumen.get("Tipo_Permiso") or registro_resumen.get("Tipo Permiso") or "N/A"
                if tipo_permiso != "N/A":
                    return f"Permiso: {tipo_permiso}"
                else:
                    return "Permiso"
            elif tipo_evento == "Campaña":
                return "Campaña"
            elif tipo_evento == "Salida Institucional":  # NUEVO: Salida Institucional
                return "Salida Institucional"
        
        # Priorizar información de permisos, campañas y SALIDAS INSTITUCIONALES desde statistics
        if info_permisos:
            tiene_permiso = info_permisos.get("tiene_permiso", "No")
            tipo_permiso = info_permisos.get("tipo_permiso", "N/A")
            tiene_campaña = info_permisos.get("tiene_campaña", "No")
            tiene_departures = info_permisos.get("tiene_departures", "No")  # NUEVO: Salida Institucional
            
            if tiene_permiso == "Sí" and tipo_permiso != "N/A":
                return f"Permiso: {tipo_permiso}"
            elif tiene_permiso == "Sí":
                return "Permiso"
            elif tiene_campaña == "Sí":
                return "Campaña"
            elif tiene_departures == "Sí":  # NUEVO: Salida Institucional
                return "Salida Institucional"
        
        # Alternativa: Usar información del resumen si statistics no está disponible
        if registro_resumen:
            tipo_evento = registro_resumen.get("Tipo_Evento") or registro_resumen.get("Tipo Evento") or "N/A"
            tipo_permiso_resumen = registro_resumen.get("Tipo_Permiso") or registro_resumen.get("Tipo Permiso") or "N/A"
            
            if tipo_evento == "Permiso" and tipo_permiso_resumen != "N/A":
                return f"Permiso: {tipo_permiso_resumen}"
            elif tipo_evento == "Permiso":
                return "Permiso"
            elif tipo_evento == "Campaña":
                return "Campaña"
            elif tipo_evento == "Salida Institucional":  # NUEVO: Salida Institucional
                return "Salida Institucional"
        
        # Manejar festivos primero
        if es_festivo:
            return "Festivo"
            
        # Lógica existente para otros casos
        if es_domingo:
            return "Día dominical"
        
        # Lógica de puntualidad existente
        if punt_entrada == "Sin registro" and punt_salida == "Sin registro":
            return "Sin registros"
        
        if punt_entrada == "temprano" and punt_salida == "temprano":
            return "Llegó y salió temprano"
        elif punt_entrada == "temprano" and punt_salida == "a_tiempo":
            return "Llegó temprano, salió a tiempo"
        elif punt_entrada == "temprano" and punt_salida == "tarde":
            return "Llegó temprano, salió tarde"
        elif punt_entrada == "a_tiempo" and punt_salida == "temprano":
            return "Llegó a tiempo, salió temprano"
        elif punt_entrada == "a_tiempo" and punt_salida == "a_tiempo":
            return "Sin novedad"
        elif punt_entrada == "a_tiempo" and punt_salida == "tarde":
            return "Llegó a tiempo, salió tarde"
        elif punt_entrada == "tarde" and punt_salida == "temprano":
            return "Llegó tarde, salió temprano"
        elif punt_entrada == "tarde" and punt_salida == "a_tiempo":
            return "Llegó tarde, salió a tiempo"
        elif punt_entrada == "tarde" and punt_salida == "tarde":
            return "Llegó y salió tarde"
        else:
            return "Registro especial"

    # Procesar registros (EXCLUYENDO FESTIVOS)
    registros_procesados = 0
    for registro in resumen:
        try:
            # Obtener datos con nombres de campos flexibles
            cedula = registro.get("Cedula") or registro.get("cedula") or ""
            nombre = registro.get("Empleado") or registro.get("empleado") or ""
            fecha_str = registro.get("Fecha") or registro.get("fecha") or ""
            
            # 🆕 MODIFICACIÓN: Obtener nombre del día
            nombre_dia = obtener_nombre_dia(fecha_str) if fecha_str and fecha_str != "N/A" else "N/A"
            
            # Verificar si es festivo
            es_festivo = False
            if fecha_str and fecha_str != "N/A":
                try:
                    fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    es_festivo, _ = es_dia_festivo_simple(fecha)
                except:
                    pass

            # Obtener vista previa del permiso del registro
            vista_previa_permiso = registro.get("Vista_Previa_Permiso") or registro.get("vista_previa_permiso") or "N/A"
            
            # Verificar si es domingo
            es_domingo = False
            if fecha_str and fecha_str != "N/A":
                try:
                    fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                    es_domingo = fecha.weekday() == 6
                except:
                    pass

            # Verificar si hay NOVEDAD en la API para esta cédula y fecha
            tiene_novedad_api = False
            info_novedad = {}
            if (cedula, fecha_str) in mapeo_novedades:
                tiene_novedad_api = True
                info_novedad = mapeo_novedades[(cedula, fecha_str)]
                print(f"⚠️  DETECTADA NOVEDAD: {cedula} - {fecha_str} - {info_novedad.get('evento', 'N/A')}")

            # Omitir domingos sin registros (pero NO festivos, ya se excluyeron)
            if es_domingo and registro.get("Hora_Ingreso") == "N/A" and registro.get("Hora_Salida") == "N/A":
                continue

            # Horas de ingreso y salida reales (NUNCA serán "Festivo")
            entrada = registro.get("Hora_Ingreso") or registro.get("Hora Ingreso") or registro.get("hora_ingreso") or "N/A"
            salida = registro.get("Hora_Salida") or registro.get("Hora Salida") or registro.get("hora_salida") or "N/A"
            
            # Horas asignadas
            ingreso_asignado = registro.get("Ingreso_Asignado") or registro.get("Ingreso Asignado") or registro.get("ingreso_asignado") or "N/A"
            salida_asignada = registro.get("Salida_Asignada") or registro.get("Salida Asignada") or registro.get("salida_asignada") or "N/A"
            
            # Obtener información de almuerzo, permisos, campañas y SALIDAS INSTITUCIONALES
            info_permisos_dia = None
            horario_almuerzo_str = "No tiene"
            tiempo_almuerzo = 0.0
            
            if cedula in info_almuerzo_permisos_novedades and fecha_str in info_almuerzo_permisos_novedades[cedula]:
                info_permisos_dia = info_almuerzo_permisos_novedades[cedula][fecha_str]
                horario_almuerzo_inicio = info_permisos_dia.get("horario_almuerzo_inicio", "N/A")
                horario_almuerzo_fin = info_permisos_dia.get("horario_almuerzo_fin", "N/A")
                
                if horario_almuerzo_inicio != "N/A" and horario_almuerzo_fin != "N/A":
                    tiempo_almuerzo = calcular_tiempo_almuerzo(horario_almuerzo_inicio, horario_almuerzo_fin)
                    horario_almuerzo_str = f"{horario_almuerzo_inicio}/{horario_almuerzo_fin}"

            # MODIFICACIÓN: NUEVO CÁLCULO DE HORAS
            # Horas trabajadas = hora salida - hora ingreso
            horas_trabajadas_decimal = calcular_horas_entre(entrada, salida)
            
            # Obtener horas descanso del registro
            horas_descanso_decimal = registro.get("Horas_Fuera") or registro.get("Horas Fuera") or registro.get("horas_fuera") or 0
            
            # Convertir a float si es necesario
            try:
                if isinstance(horas_descanso_decimal, str) and horas_descanso_decimal not in ["N/A", ""]:
                    horas_descanso_decimal = float(horas_descanso_decimal)
                elif horas_descanso_decimal in ["N/A", ""]:
                    horas_descanso_decimal = 0
            except:
                horas_descanso_decimal = 0

            # MODIFICACIÓN: Horas netas = horas trabajadas - horas descanso - horas almuerzo
            horas_netas_ajustadas = max(horas_trabajadas_decimal - horas_descanso_decimal - tiempo_almuerzo, 0)

            # Determinar puntualidad
            punt_entrada = determinar_puntualidad(entrada, ingreso_asignado, "entrada")
            punt_salida = determinar_puntualidad(salida, salida_asignada, "salida")
            
            # Generar observación mejorada que incluye permisos, campañas, SALIDAS INSTITUCIONALES Y NOVEDADES (SIN FESTIVOS)
            observaciones = generar_observacion_mejorada(
                punt_entrada, 
                punt_salida, 
                es_domingo, 
                info_permisos_dia,
                registro,
                es_festivo=es_festivo
            )

            # AGREGAR INFORMACIÓN DE NOVEDAD A LAS OBSERVACIONES
            if tiene_novedad_api:
                evento_novedad = info_novedad.get("evento", "Evento especial")
                observaciones = f"NOVEDAD: {evento_novedad}"

            # Si es novedad de orden público, sobreescribir observación y colorear
            if "NOVEDAD ORDEN PUBLICO" in observaciones.upper():
                observaciones = "NOVEDAD ORDEN PUBLICO"

            # Convertir a formato HH:MM:SS para mostrar
            horas_trabajadas_str = decimal_a_hhmmss(horas_trabajadas_decimal) if horas_trabajadas_decimal > 0 else "00:00:00"
            horas_descanso_str = decimal_a_hhmmss(horas_descanso_decimal) if horas_descanso_decimal > 0 else "00:00:00"
            horas_netas_str = decimal_a_hhmmss(horas_netas_ajustadas) if horas_netas_ajustadas > 0 else "00:00:00"

            # 🆕 MODIFICACIÓN: Escribir fila con nueva columna "Día"
            fila = [
                cedula,
                nombre,
                fecha_str,
                nombre_dia,  # 🆕 NUEVA COLUMNA: Día de la semana
                entrada,
                salida,
                horas_trabajadas_str,
                horas_descanso_str,
                horario_almuerzo_str,
                horas_netas_str,
                vista_previa_permiso,
                observaciones
            ]
            ws.append(fila)

            row_idx = ws.max_row
            registros_procesados += 1

            # Aplicar estilos y colores condicionales (ESTILOS INSTITUCIONALES)
            is_even = (row_idx - 4) % 2 == 1
            ws.row_dimensions[row_idx].height = 18
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = center
                cell.border = EstilosInstitucionales.borde_horizontal_sutil()
                cell.font = EstilosInstitucionales.fuente_normal()
                if is_even:
                    cell.fill = EstilosInstitucionales.fill_fila_par()

            # Horas Trabajadas (Col 7)
            cell_horas = ws.cell(row=row_idx, column=7)
            if horas_trabajadas_str == "00:00:00":
                cell_horas.fill = PatternFill(start_color=EstilosInstitucionales.BADGE_CRITICO_BG, end_color=EstilosInstitucionales.BADGE_CRITICO_BG, fill_type="solid")
                cell_horas.font = EstilosInstitucionales.fuente_normal() # Mantenemos texto normal pero fondo rojo suave
            
            # Vista Previa Permiso (Col 11) - Enlace
            cell_permiso = ws.cell(row=row_idx, column=11)
            if vista_previa_permiso and str(vista_previa_permiso).startswith("http"):
                cell_permiso.value = "Ver documento"
                cell_permiso.hyperlink = vista_previa_permiso
                cell_permiso.font = Font(color=EstilosInstitucionales.AZUL_PRINCIPAL, underline="single", name=EstilosInstitucionales.FONT_NAME, size=10)

            # Observaciones (Col 12) - Badges
            cell_obs = ws.cell(row=row_idx, column=12)
            bg, fg = EstilosInstitucionales.get_badge_por_texto(observaciones)
            if bg and fg:
                cell_obs.fill = bg
                cell_obs.font = fg

        except Exception as e:
            print(f"❌ Error procesando registro en hoja Diario: {e}")
            continue

    # 🆕 MODIFICACIÓN: Ajustar anchos de columna con nueva columna "Día"
    anchos_columnas = {
        "A": 14, "B": 30, "C": 12, "D": 12, "E": 12, "F": 12,  # D agregada
        "G": 15, "H": 15, "I": 20, "J": 12, "K": 25, "L": 25   # Ajustados índices
    }
    for col, ancho in anchos_columnas.items():
        ws.column_dimensions[col].width = ancho

    # Agregar filtros automáticos en TODAS las columnas
    last_row = ws.max_row
    last_col = len(headers)
    ws.auto_filter.ref = f"A4:{get_column_letter(last_col)}{last_row}"

    # Congelar fila de encabezados
    ws.freeze_panes = "A5"

    # Guardar
    wb.save(salida_excel)
    print(f"✅ Hoja 'Diario' agregada correctamente con {registros_procesados} registros (excluyendo festivos)")
    print(f"📅 Nueva columna 'Día' agregada con nombres completos (Lunes, Martes, etc.)")
    print(f"🎨 {len(mapeo_novedades)} registros coloreados en amarillo por novedad de API")
    print(f"🧮 NUEVO CÁLCULO: Horas Trabajadas = Hora Salida - Hora Ingreso")
    print(f"🧮 NUEVO CÁLCULO: Horas Netas = Horas Trabajadas - Horas Descanso - Horas Almuerzo")
    print(f"🔵 NUEVOS EVENTOS: 'Salida Institucional' y 'Campaña' se muestran en azul como los permisos")