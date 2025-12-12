# reports/excel_explicacion.py (HOJA COMPLETA DE EXPLICACIÓN)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def agregar_hoja_explicacion(salida_excel="outputs/Reporte_Asistencias_Completo.xlsx"):
    """
    Agrega una hoja de explicación completa al inicio del reporte
    """
    print("\n📖 Creando hoja de explicación...")

    try:
        # Intentar cargar el libro existente
        wb = load_workbook(salida_excel)
        # Si existe, crear la hoja al principio
        if "Explicación" in wb.sheetnames:
            del wb["Explicación"]
        ws = wb.create_sheet("Explicación", 0)  # Posición 0 = primera hoja
    except FileNotFoundError:
        # Si el archivo no existe, crear uno nuevo
        wb = Workbook()
        ws = wb.active
        ws.title = "Explicación"

    # Estilos
    title_font = Font(bold=True, size=16, color="0054A6")
    subtitle_font = Font(bold=True, size=14, color="0054A6")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    normal_font = Font(size=11)
    
    header_fill = PatternFill(start_color="0054A6", end_color="0054A6", fill_type="solid")
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    blue_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    purple_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color="E1DFDD"),
        right=Side(style='thin', color="E1DFDD"),
        top=Side(style='thin', color="E1DFDD"),
        bottom=Side(style='thin', color="E1DFDD")
    )
    
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # =========================================================================
    # TÍTULO PRINCIPAL
    # =========================================================================
    ws.merge_cells('A1:H1')
    ws['A1'] = "REPORTE DE ASISTENCIAS - GUÍA COMPLETA"
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws['A1'].fill = header_fill

    # =========================================================================
    # SECCIÓN 1: INTRODUCCIÓN
    # =========================================================================
    ws.merge_cells('A3:H3')
    ws['A3'] = "1. INTRODUCCIÓN"
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = left

    introduccion = [
        "Este sistema automatizado procesa registros de asistencia y genera reportes detallados.",
        "Combina datos de dispositivos biométricos con información de horarios programados desde la API.",
        "El análisis excluye automáticamente días festivos, domingos y empleados con menos del 10% de registros.",
        "",
        "📊 OBJETIVOS PRINCIPALES:",
        "• Analizar patrones de puntualidad y asistencia",
        "• Identificar oportunidades de mejora", 
        "• Generar métricas objetivas para la gestión del personal",
        "• Automatizar el proceso de reportes mensuales"
    ]

    for i, texto in enumerate(introduccion, 4):
        ws.merge_cells(f'A{i}:H{i}')
        ws[f'A{i}'] = texto
        ws[f'A{i}'].font = normal_font
        ws[f'A{i}'].alignment = left

    # =========================================================================
    # SECCIÓN 2: CÓDIGO DE COLORES
    # =========================================================================
    current_row = len(introduccion) + 5

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "2. CÓDIGO DE COLORES - SIGNIFICADO"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = left
    current_row += 1

    # Encabezados de tabla de colores
    headers_colores = ["Color", "Nombre", "Significado", "Dónde se usa", "Ejemplo"]
    for col, header in enumerate(headers_colores, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    current_row += 1

    # Datos de colores
    colores_data = [
        ("🟢", "Verde", "Excelente desempeño / Cumplimiento total", "Resumen Semanal, Horas trabajadas ≥8", "8+ horas trabajadas"),
        ("🟡", "Amarillo", "Desempeño regular / Necesita mejora", "Resumen Semanal, Horas trabajadas 6-8", "6-7 horas trabajadas"),
        ("🔴", "Rojo", "Desempeño deficiente / Atención requerida", "Resumen Semanal, Horas trabajadas <6", "Menos de 6 horas"),
        ("🔵", "Azul", "Permisos / Campañas / Eventos especiales", "Todas las hojas", "Días con permisos"),
        ("🟠", "Naranja", "Alerta / Preocupación", "Tabla Horarios - Estado Regular", "Puntuación 60-74%"),
        ("🟣", "Morado", "Domingos / Días no laborales", "Hoja Diario", "Registros dominicales"),
        ("⚫", "Gris", "Días festivos", "Resumen Semanal", "Festivos nacionales"),
        ("⚪", "Blanco", "Sin registros / Sin datos", "Todas las hojas", "Días sin marcación")
    ]

    for color_data in colores_data:
        for col, valor in enumerate(color_data, 1):
            cell = ws.cell(row=current_row, column=col, value=valor)
            cell.alignment = center
            cell.border = thin_border
            cell.font = normal_font
            
            # Aplicar color de fondo según el significado
            if "Verde" in valor:
                cell.fill = green_fill
            elif "Amarillo" in valor:
                cell.fill = yellow_fill
            elif "Rojo" in valor:
                cell.fill = red_fill
            elif "Azul" in valor:
                cell.fill = blue_fill
            elif "Naranja" in valor:
                cell.fill = orange_fill
            elif "Morado" in valor:
                cell.fill = purple_fill
                
        current_row += 1

    # =========================================================================
    # SECCIÓN 3: EXPLICACIÓN POR HOJAS
    # =========================================================================
    current_row += 2

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "3. EXPLICACIÓN POR HOJAS - CONTENIDO Y CÁLCULOS"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = left
    current_row += 1

    hojas_explicacion = [
        {
            "hoja": "RESUMEN SEMANAL",
            "objetivo": "Visión general semanal de horas trabajadas por empleado",
            "columnas_principales": ["Documento", "Nombre", "Cargo", "Semanas (Lun-Sáb)", "Total Permisos"],
            "calculos": "• Horas trabajadas por día • Totales semanales • Detección automática de festivos",
            "filtros": "Excluye: Empleados con <10% registros, Días sin horario programado",
            "colores": "Verde (≥8h), Amarillo (6-8h), Rojo (<6h), Azul (Permisos)"
        },
        {
            "hoja": "DIARIO", 
            "objetivo": "Detalle día a día de registros y puntualidad",
            "columnas_principales": ["Cédula", "Nombre", "Fecha", "Día", "Hora Ingreso/Salida", "Horas Trabajadas", "Observaciones"],
            "calculos": "• Puntualidad (±5 minutos) • Horas netas (restando almuerzo) • Observaciones automáticas",
            "filtros": "Excluye: Festivos, Domingos sin registros, Novedades de API",
            "colores": "Por puntualidad y tipo de novedad"
        },
        {
            "hoja": "HORARIOS",
            "objetivo": "Análisis de puntualidad y ranking de empleados",
            "columnas_principales": ["TOP 5 Tarde", "TOP 5 A Tiempo", "TOP 5 Temprano", "Lista Completa"],
            "calculos": "🔢 NUEVO SISTEMA: 40% Tarde + 35% A Tiempo + 25% Temprano = 100% Asistencia",
            "filtros": "Excluye: Docentes, Festivos, <10% registros, Días con novedad masiva",
            "colores": "Por estado general (Excelente, Bueno, Regular, A Mejorar)"
        },
        {
            "hoja": "HORARIOS PROGRAMADOS",
            "objetivo": "Visualización de horarios asignados por empleado", 
            "columnas_principales": ["Documento", "Nombre", "Semanas con días L-S"],
            "calculos": "• Horarios de entrada/salida programados • Tiempos de almuerzo",
            "filtros": "Muestra solo empleados con horario programado",
            "colores": "Sin colores - información referencial"
        },
        {
            "hoja": "REGISTROS BRUTOS",
            "objetivo": "Datos originales sin procesar para auditoría",
            "columnas_principales": "Todas las columnas originales de los archivos Excel", 
            "calculos": "Sin cálculos - datos originales",
            "filtros": "Sin filtros aplicados",
            "colores": "Sin colores - datos brutos"
        }
    ]

    for hoja in hojas_explicacion:
        # Título de la hoja
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"📊 {hoja['hoja']}"
        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="0054A6")
        ws[f'A{current_row}'].alignment = left
        ws[f'A{current_row}'].fill = PatternFill(start_color="E1DFDD", end_color="E1DFDD", fill_type="solid")
        current_row += 1

        # Información de la hoja
        info_campos = [
            f"🎯 OBJETIVO: {hoja['objetivo']}",
            f"📋 COLUMNAS: {hoja['columnas_principales']}",
            f"🧮 CÁLCULOS: {hoja['calculos']}",
            f"🔍 FILTROS: {hoja['filtros']}",
            f"🎨 COLORES: {hoja['colores']}"
        ]

        for info in info_campos:
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = info
            ws[f'A{current_row}'].font = normal_font
            ws[f'A{current_row}'].alignment = left
            current_row += 1

        current_row += 1  # Espacio entre hojas

    # =========================================================================
    # SECCIÓN 4: NUEVO SISTEMA DE CÁLCULOS - PORCENTAJES PONDERADOS
    # =========================================================================
    current_row += 1

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "4. 🔢 NUEVO SISTEMA DE CÁLCULOS - PORCENTAJES PONDERADOS"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = left
    current_row += 1

    sistema_texto = [
        "FÓRMULA GENERAL:",
        "   Puntuación = (40% × Días Tarde) + (35% × Días A Tiempo) + (25% × Días Temprano)",
        "",
        "📝 EJEMPLO PRÁCTICO:",
        "   • Días laborados totales: 20 días",
        "   • Pepito Pérez trabajó: 15 días (75% de asistencia)",
        "   • Distribución de sus 15 días:",
        "       - 3 días llegó tarde → 3 × 40% = 1.2",
        "       - 8 días llegó a tiempo → 8 × 35% = 2.8", 
        "       - 4 días llegó temprano → 4 × 25% = 1.0",
        "",
        "🧮 CÁLCULO:",
        "   Puntuación Bruta = 1.2 + 2.8 + 1.0 = 5.0",
        "   Puntuación Normalizada = (5.0 / 6.0) × 100 = 83.3%",
        "   • 6.0 es el máximo posible (15 días × 40% = 6.0)",
        "",
        "📊 INTERPRETACIÓN:",
        "   • 83.3% representa su índice de puntualidad ponderado",
        "   • Se considera en el contexto de sus días trabajados",
        "   • El sistema prioriza la consistencia sobre eventos aislados"
    ]

    for texto in sistema_texto:
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = texto
        ws[f'A{current_row}'].font = normal_font
        ws[f'A{current_row}'].alignment = left
        current_row += 1

    # =========================================================================
    # SECCIÓN 5: CRITERIOS DE CLASIFICACIÓN
    # =========================================================================
    current_row += 2

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "5. CRITERIOS DE CLASIFICACIÓN Y ESTADOS"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = left
    current_row += 1

    # Tabla de criterios
    criterios_headers = ["Estado", "Rango Puntuación", "Interpretación", "Recomendación"]
    for col, header in enumerate(criterios_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    current_row += 1

    criterios_data = [
        ("Excelente", "90-100%", "Puntualidad consistente y alta asistencia", "Mantener desempeño"),
        ("Bueno", "75-89%", "Puntualidad aceptable con oportunidades de mejora", "Reforzar consistencia"),
        ("Regular", "60-74%", "Puntualidad irregular que requiere atención", "Plan de mejora individual"),
        ("A Mejorar", "0-59%", "Puntualidad deficiente, intervención necesaria", "Acción correctiva inmediata")
    ]

    for criterio in criterios_data:
        for col, valor in enumerate(criterio, 1):
            cell = ws.cell(row=current_row, column=col, value=valor)
            cell.alignment = center
            cell.border = thin_border
            cell.font = normal_font
            
            # Colorear según estado
            if criterio[0] == "Excelente":
                cell.fill = green_fill
            elif criterio[0] == "Bueno":
                cell.fill = yellow_fill
            elif criterio[0] == "Regular":
                cell.fill = orange_fill
            else:
                cell.fill = red_fill
                
        current_row += 1

    # =========================================================================
    # SECCIÓN 6: FILTROS Y EXCLUSIONES
    # =========================================================================
    current_row += 2

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "6. FILTROS Y EXCLUSIONES AUTOMÁTICAS"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = left
    current_row += 1

    filtros_texto = [
        "🚫 EMPLEADOS EXCLUIDOS DEL ANÁLISIS:",
        "   • Docentes (no tienen horario fijo de oficina)",
        "   • Empleados con menos del 10% de registros vs días laborados",
        "   • Personal con novedades de API (permisos, campañas, etc.)",
        "",
        "📅 DÍAS EXCLUIDOS DEL ANÁLISIS:",
        "   • Días festivos nacionales",
        "   • Domingos sin registros de asistencia", 
        "   • Días con novedad masiva (>85% del personal llegó tarde)",
        "   • Días sin horario programado",
        "",
        "🔍 CRITERIOS DE PUNTUALIDAD:",
        "   • Temprano: 10+ minutos antes del horario",
        "   • A Tiempo: ±5 minutos del horario",
        "   • Tarde: 5+ minutos después del horario"
    ]

    for texto in filtros_texto:
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = texto
        ws[f'A{current_row}'].font = normal_font
        ws[f'A{current_row}'].alignment = left
        current_row += 1

    # =========================================================================
    # SECCIÓN 7: EJEMPLO COMPLETO DEL NUEVO SISTEMA
    # =========================================================================
    current_row += 2

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "7. 📊 EJEMPLO COMPLETO DEL NUEVO SISTEMA"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = left
    current_row += 1

    ejemplo_texto = [
        "EMPLEADO: María González - 20 días laborados - 18 días trabajados (90% asistencia)",
        "",
        "DISTRIBUCIÓN DE SUS 18 DÍAS:",
        "   • 5 días llegó temprano → 5 × 25% = 1.25",
        "   • 10 días llegó a tiempo → 10 × 35% = 3.50",
        "   • 3 días llegó tarde → 3 × 40% = 1.20",
        "",
        "CÁLCULO FINAL:",
        "   Puntuación Bruta = 1.25 + 3.50 + 1.20 = 5.95",
        "   Máximo Posible = 18 días × 40% = 7.20",
        "   Puntuación Final = (5.95 / 7.20) × 100 = 82.6%",
        "",
        "RESULTADO:",
        "   • Estado: BUENO (82.6% está en rango 75-89%)",
        "   • Interpretación: Desempeño aceptable con oportunidades de mejora",
        "   • Recomendación: Trabajar en reducir los días de llegada tarde"
    ]

    for texto in ejemplo_texto:
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = texto
        ws[f'A{current_row}'].font = normal_font
        ws[f'A{current_row}'].alignment = left
        current_row += 1

    # =========================================================================
    # AJUSTAR ANCHOS DE COLUMNA
    # =========================================================================
    for col in range(1, 9):  # Columnas A-H
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Guardar
    wb.save(salida_excel)
    print(f"✅ Hoja 'Explicación' creada correctamente en: {salida_excel}")

# Llamar a la función si se ejecuta directamente
if __name__ == "__main__":
    agregar_hoja_explicacion()