# reports/excel_horarios.py (VERSIÓN CORREGIDA - ALINEACIÓN Y TAMAÑOS OPTIMIZADOS)
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# Importación absoluta corregida
try:
    from reports.styles import EstilosModernos
except ImportError:
    # Clase de respaldo si no se puede importar
    class EstilosModernos:
        @staticmethod
        def fill_encabezado_principal():
            return PatternFill(start_color="0054A6", end_color="0054A6", fill_type="solid")
        @staticmethod
        def fuente_encabezado():
            return Font(bold=True, size=11, color="FFFFFF", name='Arial')
        @staticmethod
        def fill_verde():
            return PatternFill(start_color="107C10", end_color="107C10", fill_type="solid")
        @staticmethod
        def fill_naranja():
            return PatternFill(start_color="D83B01", end_color="D83B01", fill_type="solid")
        @staticmethod
        def fill_rojo():
            return PatternFill(start_color="E81123", end_color="E81123", fill_type="solid")
        @staticmethod
        def borde_sutil():
            return Border(
                left=Side(style='thin', color="E1DFDD"),
                right=Side(style='thin', color="E1DFDD"),
                top=Side(style='thin', color="E1DFDD"),
                bottom=Side(style='thin', color="E1DFDD")
            )
        @staticmethod
        def alineacion_centro():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

def agregar_hoja_horarios(resumen_path="temp/resumen_asistencias.json",
                         statistics_path="temp/statistics.json",
                         salida_excel="outputs/Reporte_Asistencias_Completo.xlsx"):
    """
    Agrega la hoja de análisis de horarios y puntualidad al reporte.
    VERSIÓN CORREGIDA: 
    - Encabezados de lista completa con ancho correcto
    - Tablas TOP 5 separadas por una columna
    - Tamaños de celdas optimizados sin espacio perdido
    """
    print("\n📘 Agregando hoja 'Horarios' (VERSIÓN CORREGIDA - ALINEACIÓN OPTIMIZADA)...")

    try:
        wb = load_workbook(salida_excel)
    except FileNotFoundError:
        print(f"❌ No se encontró el archivo {salida_excel}")
        return

    # Crear o limpiar hoja existente
    if "Horarios" in wb.sheetnames:
        ws = wb["Horarios"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Horarios")

    # Cargar datos del resumen
    try:
        with open(resumen_path, "r", encoding="utf-8") as f:
            resumen = json.load(f)
    except Exception as e:
        print(f"❌ Error al leer el archivo de resumen: {e}")
        return

    # Cargar statistics para obtener cargos y novedades
    try:
        with open(statistics_path, "r", encoding="utf-8") as f:
            statistics = json.load(f)
    except Exception as e:
        print(f"❌ Error al leer el archivo de statistics: {e}")
        statistics = []

    # Crear mapeo de cargos y novedades
    mapeo_cargos = {}
    mapeo_novedades = {}
    mapeo_permisos = {}
    
    for stat in statistics:
        try:
            cedula = str(stat.get("number_id", "")).strip()
            cargo = stat.get("cargo", "DOCENTE").strip().upper()
            año = stat.get("año")
            mes = stat.get("mes")
            dia = stat.get("dia")
            novedad = stat.get("novedad", "No")
            permiso = stat.get("permiso", "No")
            
            if cedula:
                mapeo_cargos[cedula] = cargo
                
                if permiso == "Sí":
                    if cedula not in mapeo_permisos:
                        mapeo_permisos[cedula] = 0
                    mapeo_permisos[cedula] += 1
                
                if novedad == "Sí" and año and mes and dia:
                    try:
                        fecha_str = f"{año}-{str(mes).zfill(2)}-{str(dia).zfill(2)}"
                        mapeo_novedades[(cedula, fecha_str)] = {
                            "comienzo": stat.get("novedad_comienzo", "00:00:00"),
                            "duracion": stat.get("novedad_duracion", "00:00:00"),
                            "evento": stat.get("novedad_evento", "N/A"),
                            "observaciones": stat.get("novedad_observaciones", "N/A")
                        }
                    except Exception:
                        continue
        except Exception:
            continue

    print(f"📊 Cargos cargados: {len(mapeo_cargos)} empleados")
    print(f"⚠️  Novedades detectadas: {len(mapeo_novedades)} registros")
    print(f"📋 Permisos contabilizados: {len(mapeo_permisos)} empleados con permisos")

    # Procesar datos para análisis de puntualidad (EXCLUYENDO DOCENTES)
    data_empleados = {}
    fechas_analizadas = set()
    dias_con_novedad_masiva = set()

    for registro in resumen:
        try:
            cedula = registro.get("Cedula") or registro.get("cedula") or ""
            nombre = registro.get("Empleado") or registro.get("empleado") or ""
            fecha_str = registro.get("Fecha") or registro.get("fecha") or ""
            
            # EXCLUIR DOCENTES
            cargo = mapeo_cargos.get(cedula, "DOCENTE")
            if "DOCENTE" in cargo:
                continue
            
            # EXCLUIR DÍAS CON NOVEDAD EN API
            if (cedula, fecha_str) in mapeo_novedades:
                continue
            
            ingreso_asignado = (registro.get("Ingreso_Asignado") or 
                              registro.get("Ingreso Asignado") or 
                              registro.get("ingreso_asignado") or 
                              registro.get("IngresoAsignado") or "N/A")
            
            hora_ingreso = (registro.get("Hora_Ingreso") or 
                          registro.get("Hora Ingreso") or 
                          registro.get("hora_ingreso") or 
                          registro.get("HoraIngreso") or "N/A")
            
            hora_salida = (registro.get("Hora_Salida") or 
                         registro.get("Hora Salida") or 
                         registro.get("hora_salida") or 
                         registro.get("HoraSalida") or "N/A")
            
            salida_asignada = (registro.get("Salida_Asignada") or 
                             registro.get("Salida Asignada") or 
                             registro.get("salida_asignada") or 
                             registro.get("SalidaAsignada") or "N/A")

            if (ingreso_asignado == "N/A" or hora_ingreso == "N/A" or 
                not fecha_str or fecha_str == "N/A"):
                continue

            try:
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except:
                try:
                    fecha = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                except:
                    continue

            # EXCLUIR DOMINGOS
            if fecha.weekday() == 6:
                continue

            try:
                ingreso_programado = datetime.strptime(ingreso_asignado, "%H:%M:%S")
                ingreso_real = datetime.strptime(hora_ingreso, "%H:%M:%S")
                
                minutos_ingreso_favor = (ingreso_programado - ingreso_real).total_seconds() / 60
                minutos_ingreso_favor = max(minutos_ingreso_favor, 0)
                
                minutos_salida_favor = 0
                if hora_salida != "N/A" and salida_asignada != "N/A":
                    salida_programada = datetime.strptime(salida_asignada, "%H:%M:%S")
                    salida_real = datetime.strptime(hora_salida, "%H:%M:%S")
                    minutos_salida_favor = (salida_real - salida_programada).total_seconds() / 60
                    minutos_salida_favor = max(minutos_salida_favor, 0)
                
                minutos_totales_favor = minutos_ingreso_favor + minutos_salida_favor
                
            except:
                continue

            fechas_analizadas.add(fecha)

            if cedula not in data_empleados:
                data_empleados[cedula] = {
                    "nombre": nombre,
                    "cargo": cargo,
                    "fechas_temprano": set(),
                    "fechas_atiempo": set(),
                    "fechas_tarde": set(),
                    "fechas_analizadas": set(),
                    "minutos_retraso_acumulados": 0,
                    "minutos_favor_acumulados": 0
                }

            # Calcular puntualidad
            minutos_diferencia = (ingreso_real - ingreso_programado).total_seconds() / 60
            
            data_empleados[cedula]["fechas_analizadas"].add(fecha)
            data_empleados[cedula]["minutos_favor_acumulados"] += minutos_totales_favor

            if minutos_diferencia <= -10:  # 10+ minutos temprano
                data_empleados[cedula]["fechas_temprano"].add(fecha)
            elif -10 < minutos_diferencia <= 5:  # Entre -10 y +5 minutos (a tiempo)
                data_empleados[cedula]["fechas_atiempo"].add(fecha)
            else:  # Más de 5 minutos tarde
                data_empleados[cedula]["fechas_tarde"].add(fecha)
                data_empleados[cedula]["minutos_retraso_acumulados"] += minutos_diferencia

        except Exception as e:
            continue

    # Detectar días con novedad masiva (>85% llegaron tarde)
    for fecha in fechas_analizadas:
        total_empleados_dia = 0
        empleados_tarde_dia = 0
        
        for cedula, info in data_empleados.items():
            if fecha in info["fechas_analizadas"]:
                total_empleados_dia += 1
                if fecha in info["fechas_tarde"]:
                    empleados_tarde_dia += 1
        
        if total_empleados_dia > 0:
            porcentaje_tarde = (empleados_tarde_dia / total_empleados_dia) * 100
            if porcentaje_tarde >= 85:
                dias_con_novedad_masiva.add(fecha)
                print(f"⚠️  DETECTADA NOVEDAD MASIVA: {fecha} - {porcentaje_tarde:.1f}% llegaron tarde")

    # Excluir días con novedad masiva del análisis
    for cedula, info in data_empleados.items():
        for fecha in dias_con_novedad_masiva:
            if fecha in info["fechas_analizadas"]:
                info["fechas_analizadas"].discard(fecha)
                info["fechas_temprano"].discard(fecha)
                info["fechas_atiempo"].discard(fecha)
                info["fechas_tarde"].discard(fecha)

    if not fechas_analizadas:
        print("❌ No hay fechas válidas para analizar")
        ws.append(["ANÁLISIS DE HORARIOS"])
        ws.append(["No se encontraron datos suficientes para el análisis de puntualidad"])
        wb.save(salida_excel)
        return

    # Calcular período analizado
    fecha_min = min(fechas_analizadas) if fechas_analizadas else None
    fecha_max = max(fechas_analizadas) if fechas_analizadas else None
    periodo_analizado = f"{fecha_min.strftime('%d/%m/%Y')} - {fecha_max.strftime('%d/%m/%Y')}" if fecha_min and fecha_max else "Período no disponible"

    # Detectar días con baja asistencia (≥80% NO asistieron)
    dias_con_bajo_porcentaje_asistencia = set()
    registros_por_fecha = {}
    
    for registro in resumen:
        try:
            cedula = registro.get("Cedula") or registro.get("cedula") or ""
            fecha_str = registro.get("Fecha") or registro.get("fecha") or ""
            
            cargo = mapeo_cargos.get(cedula, "DOCENTE")
            if "DOCENTE" in cargo:
                continue
            
            if not fecha_str or fecha_str == "N/A":
                continue
            
            try:
                fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
            except:
                try:
                    fecha = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                except:
                    continue
            
            if fecha.weekday() == 6:
                continue
            
            if fecha not in registros_por_fecha:
                registros_por_fecha[fecha] = {"total": 0, "no_asistieron": 0}
            
            registros_por_fecha[fecha]["total"] += 1
            
            hora_ingreso = (registro.get("Hora_Ingreso") or 
                          registro.get("Hora Ingreso") or 
                          registro.get("hora_ingreso") or 
                          registro.get("HoraIngreso") or "N/A")
            
            if hora_ingreso == "N/A":
                registros_por_fecha[fecha]["no_asistieron"] += 1
        except:
            continue
    
    for fecha, stats in registros_por_fecha.items():
        if stats["total"] > 0:
            porcentaje_no_asistieron = (stats["no_asistieron"] / stats["total"]) * 100
            if porcentaje_no_asistieron >= 80:
                dias_con_bajo_porcentaje_asistencia.add(fecha)
                print(f"⚠️  DETECTADO DÍA CON BAJA ASISTENCIA: {fecha} - {porcentaje_no_asistieron:.1f}% NO asistieron")
    
    # Calcular total de días laborales en el período
    total_dias_laborales_periodo = 0
    if fecha_min and fecha_max:
        current_date = fecha_min
        while current_date <= fecha_max:
            if current_date.weekday() != 6 and current_date not in dias_con_bajo_porcentaje_asistencia:
                total_dias_laborales_periodo += 1
            current_date += timedelta(days=1)

    print(f"📊 Análisis de horarios - EMPLEADOS NO DOCENTES: {len(data_empleados)}")
    print(f"📅 Período analizado: {periodo_analizado}")
    print(f"📅 Total días laborales período: {total_dias_laborales_periodo}")
    print(f"⚠️  Días con novedad masiva excluidos: {len(dias_con_novedad_masiva)}")

    # ============================================================================
    # ESTILOS UNIFICADOS CON HOJA DE PERMISOS
    # ============================================================================
    
    # Colores para fondos (igual que hoja de permisos)
    header_fill = PatternFill(start_color="0054A6", end_color="0054A6", fill_type="solid")  # Azul principal
    subheader_fill_izquierda = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")  # Azul para tabla izquierda
    subheader_fill_derecha = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Azul más oscuro para tabla derecha
    
    # Fuentes (igual que hoja de permisos)
    header_font_large = Font(bold=True, size=14, color="FFFFFF", name='Arial')
    header_font_medium = Font(bold=True, size=11, color="FFFFFF", name='Arial')
    header_font_small = Font(bold=True, size=10, color="FFFFFF", name='Arial')
    normal_font = Font(size=10, name='Arial')
    small_font = Font(size=8, name='Arial')  # Fuente más pequeña para días
    
    # Bordes (igual que hoja de permisos)
    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"),
        right=Side(style='thin', color="BFBFBF"),
        top=Side(style='thin', color="BFBFBF"),
        bottom=Side(style='thin', color="BFBFBF")
    )
    
    medium_border = Border(
        left=Side(style='medium', color="0054A6"),
        right=Side(style='medium', color="0054A6"),
        top=Side(style='medium', color="0054A6"),
        bottom=Side(style='medium', color="0054A6")
    )
    
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Colores para estados (CORREGIDOS - DIFERENTES PARA REGULAR Y A MEJORAR)
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")  # Verde
    yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Amarillo
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Naranja (REGULAR)
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Rojo (A MEJORAR)

    # FUNCIÓN PARA CALCULAR PUNTUACIÓN COHERENTE
    def calcular_puntuacion_coherente(dias_temprano, dias_atiempo, dias_tarde, dias_analizados):
        if dias_analizados == 0:
            return 0, 0, 0, 0
            
        porcentaje_temprano = (dias_temprano / dias_analizados) * 100
        porcentaje_atiempo = (dias_atiempo / dias_analizados) * 100
        porcentaje_tarde = (dias_tarde / dias_analizados) * 100
        
        # FÓRMULA COHERENTE: (a_tiempo + temprano - tarde)
        puntuacion = porcentaje_atiempo + porcentaje_temprano - porcentaje_tarde
        puntuacion = max(0, min(100, puntuacion))  # Limitar entre 0-100
        
        return porcentaje_temprano, porcentaje_atiempo, porcentaje_tarde, puntuacion

    # PREPARAR DATOS COHERENTES PARA TODAS LAS TABLAS
    empleados_para_analisis = []
    
    for cedula, info in data_empleados.items():
        dias_analizados = len(info["fechas_analizadas"])
        
        # FILTRO: Mínimo 50% de días analizados del período total
        porcentaje_asistencia = (dias_analizados * 100) / total_dias_laborales_periodo
        if porcentaje_asistencia < 50:
            continue
            
        dias_temprano = len(info["fechas_temprano"])
        dias_atiempo = len(info["fechas_atiempo"])
        dias_tarde = len(info["fechas_tarde"])
        
        # CALCULAR PORCENTAJES Y PUNTUACIÓN (COHERENTE)
        pct_temprano, pct_atiempo, pct_tarde, puntuacion = calcular_puntuacion_coherente(
            dias_temprano, dias_atiempo, dias_tarde, dias_analizados
        )
        
        empleados_para_analisis.append({
            "cedula": cedula,
            "nombre": info["nombre"],
            "dias_analizados": dias_analizados,
            "dias_temprano": dias_temprano,
            "dias_atiempo": dias_atiempo,
            "dias_tarde": dias_tarde,
            "pct_temprano": pct_temprano,
            "pct_atiempo": pct_atiempo,
            "pct_tarde": pct_tarde,
            "puntuacion": puntuacion,
            "minutos_retraso": round(info["minutos_retraso_acumulados"]),
            "minutos_favor": round(info["minutos_favor_acumulados"]),
            "permisos": mapeo_permisos.get(cedula, 0)
        })

    # ============================================================================
    # 🎯 NUEVA LÓGICA: TOP 5 COHERENTES CON LISTA COMPLETA
    # ============================================================================

    # Ordenar por puntuación (DESCENDENTE) - COHERENTE CON LA FÓRMULA
    empleados_ordenados = sorted(empleados_para_analisis, key=lambda x: -x["puntuacion"])

    # DEFINIR LOS TOP 5 BASADOS EN CRITERIOS ESPECÍFICOS
    top5_mejor_puntuacion = empleados_ordenados[:5]  # Primeros 5 de la lista completa (mejor puntuación)
    
    # TOP 5 CON MÁS TARDANZAS: Ordenar por porcentaje de tardanza descendente
    top5_mayor_tardanza = sorted(empleados_para_analisis, key=lambda x: -x["pct_tarde"])[:5]

    print(f"🎯 TOP 5 Mejor Puntuación: {len(top5_mejor_puntuacion)} empleados")
    print(f"🎯 TOP 5 Mayor Tardanza: {len(top5_mayor_tardanza)} empleados")

    # ============================================================================
    # TABLAS TOP 5 SEPARADAS POR COLUMNA (VERSIÓN OPTIMIZADA)
    # ============================================================================

    # Títulos generales - ANCHO DE 20 COLUMNAS (9 + 1 + 9 + 1)
    titulos_generales = [
        "ANÁLISIS DE PUNTUALIDAD - EMPLEADOS NO DOCENTES",
        f"Período analizado: {periodo_analizado}",
        f"Total días laborales en período: {total_dias_laborales_periodo}",
        f"Días excluidos por novedad masiva: {len(dias_con_novedad_masiva)}",
        "Mínimo 50% de días analizados del total del período"
    ]

    for titulo in titulos_generales:
        ws.append([titulo])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = header_font_medium
        cell.fill = header_fill
        cell.alignment = center
        cell.border = medium_border
        # ANCHO DE 20 COLUMNAS (9 + 1 separación + 9 + 1 margen)
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=20)

    ws.append([])

    # =====================================================================
    # TABLAS TOP 5 SEPARADAS POR UNA COLUMNA
    # =====================================================================

    # Títulos de las tablas TOP 5 (separadas por una columna)
    titulo_izquierdo = "TOP 5 - EMPLEADOS CON MEJOR PUNTUALIDAD (+5 MIN ANTES HORARIO Y MENOS % TARDANZAS)"
    titulo_derecho = "TOP 5 - EMPLEADOS CON MÁS TARDANZAS MAS % DE TARDANZAS"
    
    # Fila de títulos: tabla izquierda (cols 1-9), separación (col 10), tabla derecha (cols 11-19), margen (col 20)
    ws.append([titulo_izquierdo] + [""] * 8 + [""] + [titulo_derecho] + [""] * 8 + [""])

    # Estilos de los títulos superiores
    row_title = ws.max_row
    ws.merge_cells(start_row=row_title, start_column=1, end_row=row_title, end_column=9)   # Tabla izquierda
    ws.merge_cells(start_row=row_title, start_column=11, end_row=row_title, end_column=19) # Tabla derecha

    # Aplicar colores diferentes a los títulos
    cell_izquierda = ws.cell(row=row_title, column=1)
    cell_izquierda.font = header_font_medium
    cell_izquierda.fill = subheader_fill_izquierda
    cell_izquierda.alignment = center
    cell_izquierda.border = medium_border

    cell_derecha = ws.cell(row=row_title, column=11)
    cell_derecha.font = header_font_medium
    cell_derecha.fill = subheader_fill_derecha
    cell_derecha.alignment = center
    cell_derecha.border = medium_border

    # ENCABEZADOS - MISMA ESTRUCTURA PARA AMBAS TABLAS
    subtitulos_izquierda = ["#", "Documento", "Nombre", "D Temprano", "D Tiempo", "D Tarde", "D Analizados", "Puntaje", "Estado"]
    subtitulos_derecha = ["#", "Documento", "Nombre", "D Tarde", "D Analizados", "% Tarde", "Estado", "", ""]  # 2 vacías para igualar

    # Fila de encabezados: tabla izquierda + separación + tabla derecha + margen
    ws.append(subtitulos_izquierda + [""] + subtitulos_derecha + [""])

    # Estilos encabezados
    row_headers = ws.max_row
    for col in range(1, 21):
        cell = ws.cell(row=row_headers, column=col)
        cell.alignment = center
        cell.border = thin_border
        
        if col <= 9:  # Tabla izquierda
            cell.font = header_font_small
            cell.fill = subheader_fill_izquierda
        elif col == 10:  # Separación
            cell.fill = PatternFill(fill_type=None)  # Sin relleno
        elif col <= 19:  # Tabla derecha
            cell.font = header_font_small
            cell.fill = subheader_fill_derecha
        else:  # Margen
            cell.fill = PatternFill(fill_type=None)

    # =====================================================================
    # ESCRIBIR FILAS DE TOP 5 (SEPARADAS POR COLUMNA)
    # =====================================================================

    for i in range(5):
        fila = []

        # -----------------------
        # 🟩 TABLA IZQUIERDA (MEJOR PUNTUALIDAD)
        # -----------------------
        if i < len(top5_mejor_puntuacion):
            emp = top5_mejor_puntuacion[i]

            # Estado basado en puntuación general
            if emp["puntuacion"] >= 80:
                estado = "Excelente"
            elif emp["puntuacion"] >= 60:
                estado = "Bueno"
            elif emp["puntuacion"] >= 40:
                estado = "Regular"
            else:
                estado = "A Mejorar"

            fila += [
                i + 1,
                emp["cedula"],
                emp["nombre"],
                f"{emp['dias_temprano']}",
                f"{emp['dias_atiempo']}",
                f"{emp['dias_tarde']}",
                emp["dias_analizados"],
                f"{emp['puntuacion']:.1f}%",
                estado
            ]
        else:
            fila += [""] * 9

        # -----------------------
        # SEPARACIÓN (COLUMNA 10)
        # -----------------------
        fila.append("")

        # -----------------------
        # 🟥 TABLA DERECHA (MÁS TARDANZAS)
        # -----------------------
        if i < len(top5_mayor_tardanza):
            emp_tarde = top5_mayor_tardanza[i]

            # Estado basado en porcentaje de tardanza
            if emp_tarde["pct_tarde"] >= 50:
                estado_t = "Crítico"
            elif emp_tarde["pct_tarde"] >= 30:
                estado_t = "Alto"
            elif emp_tarde["pct_tarde"] >= 20:
                estado_t = "Medio"
            else:
                estado_t = "Bajo"

            fila += [
                i + 1,
                emp_tarde["cedula"],
                emp_tarde["nombre"],
                f"{emp_tarde['dias_tarde']}",
                emp_tarde["dias_analizados"],
                f"{emp_tarde['pct_tarde']:.1f}%",
                estado_t,
                "",  # Columna vacía
                ""   # Columna vacía
            ]
        else:
            fila += [""] * 9

        # Margen final (columna 20)
        fila.append("")

        ws.append(fila)

        # Aplicar estilos a la fila completa
        current_row = ws.max_row
        for col in range(1, 21):
            c = ws.cell(row=current_row, column=col)
            c.border = thin_border
            c.alignment = center
            
            # Aplicar fuente pequeña a las columnas de días
            if col in [4, 5, 6, 7, 14, 15]:
                c.font = small_font
            else:
                c.font = normal_font

        # Colores de estados TABLA IZQUIERDA
        if i < len(top5_mejor_puntuacion):
            estado_cell_left = ws.cell(row=current_row, column=9)
            if estado == "Excelente":
                estado_cell_left.fill = green_fill
            elif estado == "Bueno":
                estado_cell_left.fill = yellow_fill
            elif estado == "Regular":
                estado_cell_left.fill = orange_fill
            else:
                estado_cell_left.fill = red_fill

        # Colores de estados TABLA DERECHA
        if i < len(top5_mayor_tardanza):
            estado_cell_right = ws.cell(row=current_row, column=16)
            if estado_t == "Crítico":
                estado_cell_right.fill = red_fill
            elif estado_t == "Alto":
                estado_cell_right.fill = orange_fill
            elif estado_t == "Medio":
                estado_cell_right.fill = yellow_fill
            else:
                estado_cell_right.fill = green_fill

    # ============================================================================
    # LISTA COMPLETA (ANCHO CORRECTO - 11 COLUMNAS)
    # ============================================================================

    ws.append([])
    ws.append([])
    
    # Título lista completa - ANCHO CORRECTO (11 columnas)
    titulos_completa = [
        "LISTA COMPLETA - PUNTUALIDAD GENERAL (EMPLEADOS NO DOCENTES)",
        f"Período analizado: {periodo_analizado}",
        f"Total de empleados analizados: {len(empleados_para_analisis)} | Días laborales en período: {total_dias_laborales_periodo}",
        f"Días excluidos por novedad masiva: {len(dias_con_novedad_masiva)}",
        "FÓRMULA: Puntuación = (%A_Tiempo + %Temprano - %Tarde)",
        "ORDEN: Por Puntuación (DESCENDENTE)",
        "NOTA: Los TOP 5 están extraídos de esta lista completa"
    ]

    for titulo in titulos_completa:
        ws.append([titulo])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = header_font_medium
        cell.fill = header_fill
        cell.alignment = center
        cell.border = medium_border
        # ANCHO CORRECTO: 11 columnas para la lista completa
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=11)

    ws.append([])

    # Encabezados lista completa - ANCHO CORRECTO (11 columnas)
    headers_completa = ["#", "Documento", "Nombre", "D Laborados", "D Temprano", "D Tiempo", "D Tarde", "Min Retraso", "Min Favor", "Permisos", "Estado"]
    ws.append(headers_completa)
    for col_num, header in enumerate(headers_completa, 1):
        cell = ws.cell(row=ws.max_row, column=col_num)
        cell.font = header_font_small
        cell.fill = subheader_fill_izquierda
        cell.alignment = center
        cell.border = medium_border

    # Escribir lista completa - SOLO 11 COLUMNAS
    for idx, emp in enumerate(empleados_ordenados, 1):
        # Determinar estado general
        if emp["puntuacion"] >= 80:
            estado_general = "Excelente"
        elif emp["puntuacion"] >= 60:
            estado_general = "Bueno"
        elif emp["puntuacion"] >= 40:
            estado_general = "Regular"
        else:
            estado_general = "A Mejorar"

        ws.append([
            idx,
            emp["cedula"],
            emp["nombre"],
            emp["dias_analizados"],
            f"{emp['dias_temprano']} ({emp['pct_temprano']:.1f}%)",
            f"{emp['dias_atiempo']} ({emp['pct_atiempo']:.1f}%)",
            f"{emp['dias_tarde']} ({emp['pct_tarde']:.1f}%)",
            f"{emp['minutos_retraso']}",
            f"{emp['minutos_favor']}",
            emp["permisos"],
            f"{estado_general} ({emp['puntuacion']:.1f}%)"
        ])
        
        current_row = ws.max_row
        for col in range(1, 12):  # Solo hasta la columna K (11)
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = center
            
            # Aplicar fuente pequeña a las columnas de días
            if col in [4, 5, 6, 7]:
                cell.font = small_font
            else:
                cell.font = normal_font
            
        # Colorear estado general
        estado_cell = ws.cell(row=current_row, column=11)
        if estado_general == "Excelente":
            estado_cell.fill = green_fill
        elif estado_general == "Bueno":
            estado_cell.fill = yellow_fill
        elif estado_general == "Regular":
            estado_cell.fill = orange_fill
        else:
            estado_cell.fill = red_fill

    # ============================================================================
    # AJUSTES FINALES DE TAMAÑOS DE COLUMNAS (OPTIMIZADOS)
    # ============================================================================

    # Anchuras optimizadas para evitar espacio perdido
    anchos_columnas = {
        # Tablas TOP 5 (columnas 1-19)
        1: 4,   2: 12,  3: 20,  4: 8,   5: 8,   6: 8,   7: 8,   8: 8,   9: 10,   # Tabla izquierda
        10: 2,  # Separación
        11: 4,  12: 12, 13: 20, 14: 8,  15: 8,  16: 8,  17: 10, 18: 5,  19: 5,   # Tabla derecha
        20: 2,  # Margen
        
        # Lista completa (columnas A-K)
        'A': 4,   'B': 12,  'C': 20,  'D': 10,  'E': 10,  'F': 10,  
        'G': 10,  'H': 10,  'I': 10,  'J': 8,   'K': 12
    }
    
    for col, ancho in anchos_columnas.items():
        if isinstance(col, int):
            ws.column_dimensions[chr(64 + col)].width = ancho
        else:
            ws.column_dimensions[col].width = ancho

    # Guardar
    wb.save(salida_excel)
    print(f"✅ Hoja 'Horarios' agregada correctamente")
    print(f"📊 Período analizado: {periodo_analizado}")
    print(f"👥 Empleados no docentes analizados: {len(empleados_para_analisis)}")
    print(f"🎯 TOP 5 Mejor Puntuación: Primeros 5 de lista completa")
    print(f"🎯 TOP 5 Más Tardanzas: Ordenados por % de tardanza (más impuntual primero)")
    print("📐 ALINEACIÓN CORREGIDA:")
    print("   - Encabezados lista completa con ancho correcto (11 columnas)")
    print("   - Tablas TOP 5 separadas por una columna")
    print("   - Tamaños de celdas optimizados sin espacio perdido")
    print("   - Formato compacto y profesional")