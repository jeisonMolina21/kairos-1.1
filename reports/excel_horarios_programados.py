# reports/excel_horarios_programados.py (MODIFICADO)
import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
import os

# Agregar el directorio raíz al path para importaciones absolutas
current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, current_dir)

# Importación robusta de EstilosModernos
try:
    from reports.styles import EstilosModernos
    print("✅ EstilosModernos importado correctamente desde reports.styles")
except ImportError as e:
    print(f"⚠️ No se pudo importar EstilosModernos: {e}")
    # Clase de respaldo
    class EstilosModernos:
        AZUL_PRINCIPAL = "0054A6"
        VERDE_SUTIL = "C6E0B4"
        AMARILLO_SUTIL = "FFE699"
        ROJO_SUTIL = "F8CBAD"
        
        @staticmethod
        def fill_encabezado_principal():
            return PatternFill(start_color=EstilosModernos.AZUL_PRINCIPAL, 
                              end_color=EstilosModernos.AZUL_PRINCIPAL, 
                              fill_type="solid")
        
        @staticmethod
        def fuente_encabezado():
            return Font(bold=True, size=11, color="FFFFFF", name='Arial')
        
        @staticmethod
        def fill_verde_sutil():
            return PatternFill(start_color=EstilosModernos.VERDE_SUTIL, 
                              end_color=EstilosModernos.VERDE_SUTIL, 
                              fill_type="solid")
        
        @staticmethod
        def fill_amarillo_sutil():
            return PatternFill(start_color=EstilosModernos.AMARILLO_SUTIL, 
                              end_color=EstilosModernos.AMARILLO_SUTIL, 
                              fill_type="solid")
        
        @staticmethod
        def fill_rojo_sutil():
            return PatternFill(start_color=EstilosModernos.ROJO_SUTIL, 
                              end_color=EstilosModernos.ROJO_SUTIL, 
                              fill_type="solid")
        
        @staticmethod
        def borde_sutil():
            return Border(
                left=Side(style='thin', color="E1DFDD"),
                right=Side(style='thin', color="E1DFDD"),
                top=Side(style='thin', color="E1DFDD"),
                bottom=Side(style='thin', color="E1DFDD")
            )
        
        @staticmethod
        def alineacion_centro():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

# Función simple para festivos (para excluirlos)
def es_dia_festivo_simple(fecha):
    """
    Función simplificada para verificar días festivos.
    """
    festivos_colombia = {
        '01-01': 'Año Nuevo',
        '01-06': 'Día de los Reyes Magos',
        '03-25': 'Día de San José',
        '04-13': 'Jueves Santo',
        '04-14': 'Viernes Santo',
        '05-01': 'Día del Trabajo',
        '05-22': 'Día de la Ascensión',
        '06-12': 'Corpus Christi',
        '06-19': 'Sagrado Corazón',
        '07-03': 'San Pedro y San Pablo',
        '07-20': 'Día de la Independencia',
        '08-07': 'Batalla de Boyacá',
        '08-21': 'Asunción de la Virgen',
        '10-16': 'Día de la Raza',
        '11-06': 'Todos los Santos',
        '11-13': 'Independencia de Cartagena',
        '12-08': 'Día de la Inmaculada Concepción',
        '12-25': 'Navidad'
    }
    
    fecha_str = fecha.strftime('%m-%d')
    if fecha_str in festivos_colombia:
        return True, festivos_colombia[fecha_str]
    return False, None

def agregar_hoja_horarios_programados(statistics_path="outputs/statistics.json",
                                     salida_excel="outputs/Reporte_Asistencias_Completo.xlsx"):
    """
    Agrega la hoja de horarios programados al reporte con estructura similar al resumen semanal.
    """
    print("\n📘 Agregando hoja 'Horarios Programados'...")

    try:
        wb = load_workbook(salida_excel)
    except FileNotFoundError:
        print(f"❌ No se encontró el archivo {salida_excel}")
        return

    # Crear o limpiar hoja existente
    if "Horarios Programados" in wb.sheetnames:
        ws = wb["Horarios Programados"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Horarios Programados")

    # Cargar datos
    try:
        with open(statistics_path, "r", encoding="utf-8") as f:
            statistics = json.load(f)
    except Exception as e:
        print(f"❌ Error al leer el archivo de statistics: {e}")
        return

    if not statistics:
        print("❌ No hay datos para generar horarios programados")
        return

    # Convertir a DataFrame para procesamiento
    df = pd.DataFrame(statistics)
    
    # Normalizar columnas
    df.columns = [col.strip().title() for col in df.columns]
    
    # Crear columna de fecha
    df["Fecha"] = pd.to_datetime(df[["Año", "Mes", "Dia"]].rename(columns={
        "Año": "year", "Mes": "month", "Dia": "day"
    }), errors="coerce")
    
    df = df.dropna(subset=["Fecha"])
    
    # Normalizar IDs
    df["Cedula"] = df.get("Number_Id", pd.Series([""]*len(df))).fillna("").astype(str)

    # Determinar rango de fechas
    if df.empty:
        print("❌ DataFrame vacío después de procesar")
        return
        
    fecha_inicio = df["Fecha"].min()
    fecha_fin = df["Fecha"].max()
    print(f"🗓️ Rango detectado para horarios: {fecha_inicio.date()} → {fecha_fin.date()}")

    # Obtener festivos para los años en el rango usando la función simple
    festivos = []
    for año in range(fecha_inicio.year, fecha_fin.year + 1):
        # Generar todas las fechas del año y verificar si son festivos
        fecha_actual = datetime(año, 1, 1).date()
        while fecha_actual.year == año:
            es_festivo, _ = es_dia_festivo_simple(fecha_actual)
            if es_festivo:
                festivos.append(fecha_actual)
            fecha_actual += timedelta(days=1)

    # Estilos modernos
    header_fill = EstilosModernos.fill_encabezado_principal()
    header_font = EstilosModernos.fuente_encabezado()
    center = EstilosModernos.alineacion_centro()
    border = EstilosModernos.borde_sutil()

    # Encabezados fijos
    headers_fijos = ["Documento", "Nombre"]
    for col_idx, header in enumerate(headers_fijos, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    # Encabezados dinámicos por semana - SIN columna "Horas Semana"
    current_col = len(headers_fijos) + 1
    dias_semana = ["L", "M", "X", "J", "V", "S"]  # Removido "Horas Semana"
    
    # Calcular semanas en el rango
    semanas = []
    inicio_semana = fecha_inicio - timedelta(days=fecha_inicio.weekday())
    
    while inicio_semana <= fecha_fin:
        fin_semana = inicio_semana + timedelta(days=6)
        semanas.append((inicio_semana, fin_semana))
        inicio_semana = fin_semana + timedelta(days=1)

    # Escribir encabezados de semanas
    for i, (inicio, fin) in enumerate(semanas, 1):
        col_inicio = get_column_letter(current_col)
        col_fin = get_column_letter(current_col + 5)  # 6 días ahora
        
        # Encabezado principal de la semana
        ws.merge_cells(f"{col_inicio}1:{col_fin}1")
        cell = ws[f"{col_inicio}1"]
        cell.value = f"Semana {i}\n{inicio.strftime('%d/%m')} - {fin.strftime('%d/%m')}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

        # Días de la semana
        for j, dia in enumerate(dias_semana):
            cell = ws.cell(row=2, column=current_col + j, value=dia)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        current_col += 6  # 6 columnas por semana ahora

    # Rellenar datos
    fila_actual = 3
    empleados = df["Nombre"].unique()

    for empleado in empleados:
        data_emp = df[df["Nombre"] == empleado]
        cedula = data_emp["Cedula"].iloc[0] if not data_emp.empty else ""
        
        # Datos del empleado
        ws.cell(row=fila_actual, column=1, value=cedula).border = border
        ws.cell(row=fila_actual, column=2, value=empleado).border = border

        current_col = 3
        
        # Procesar cada semana
        for inicio, fin in semanas:
            # Procesar cada día de la semana
            for j in range(6):  # Lunes a Sábado
                fecha_dia = inicio + timedelta(days=j)
                
                # Verificar si es festivo
                es_festivo = fecha_dia.date() in festivos
                
                registros_dia = data_emp[data_emp["Fecha"].dt.date == fecha_dia.date()]
                
                valor = ""
                
                if es_festivo:
                    valor = "Festivo"
                elif not registros_dia.empty:
                    registro = registros_dia.iloc[0]
                    horario_entrada = registro.get("Horario_Entrada", "N/A")
                    horario_salida = registro.get("Horario_Salida", "N/A")
                    horario_almuerzo_inicio = registro.get("Horario_Almuerzo_Inicio", "N/A")
                    horario_almuerzo_fin = registro.get("Horario_Almuerzo_Fin", "N/A")
                    tiene_horario = registro.get("Horario", "No") == "Sí"
                    
                    if tiene_horario and horario_entrada != "N/A" and horario_salida != "N/A":
                        # Calcular tiempo de almuerzo
                        tiempo_almuerzo = "No tiene"
                        if horario_almuerzo_inicio != "N/A" and horario_almuerzo_fin != "N/A":
                            try:
                                h_almuerzo_inicio = datetime.strptime(horario_almuerzo_inicio, "%H:%M:%S")
                                h_almuerzo_fin = datetime.strptime(horario_almuerzo_fin, "%H:%M:%S")
                                duracion_almuerzo = (h_almuerzo_fin - h_almuerzo_inicio).total_seconds() / 3600
                                tiempo_almuerzo = f"Almuerzo {duracion_almuerzo:.1f}h"
                            except:
                                tiempo_almuerzo = "Error almuerzo"
                        
                        valor = f"{horario_entrada}/{horario_salida}\n{tiempo_almuerzo}"
                    else:
                        valor = "Sin horario"
                else:
                    valor = "No programado"
                
                # Escribir celda - SIN color de fondo
                cell = ws.cell(row=fila_actual, column=current_col + j, value=valor)
                cell.alignment = center
                cell.border = border

            current_col += 6  # 6 columnas por semana

        fila_actual += 1

    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    for col in range(3, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # MODIFICACIÓN: Agregar filtros automáticos SOLO en documento y nombre
    ws.auto_filter.ref = f"A1:B{fila_actual-1}"

    # MODIFICACIÓN: Congelar paneles para que se vean los encabezados al bajar
    ws.freeze_panes = "A3"

    # Guardar
    wb.save(salida_excel)
    print(f"✅ Hoja 'Horarios Programados' creada en: {salida_excel}")
    print("🔍 FILTROS: Ahora solo en Documento y Nombre")