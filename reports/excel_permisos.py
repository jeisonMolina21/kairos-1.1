# reports/excel_permisos.py (MODIFICADO - MEJOR DISEÑO VISUAL)
import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys
import os

# Agregar el directorio raíz al path para importaciones absolutas
current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, current_dir)

# Importación robusta de EstilosModernos
try:
    from reports.styles import EstilosModernos
    print("✅ EstilosModernos importado correctamente desde reports.styles")
except ImportError as e:
    print(f"⚠️ No se pudo importar EstilosModernos: {e}")
    # Clase de respaldo
    class EstilosModernos:
        AZUL_PRINCIPAL = "0054A6"
        VERDE_SUTIL = "C6E0B4"
        AMARILLO_SUTIL = "FFE699"
        ROJO_SUTIL = "F8CBAD"
        GRIS_CLARO = "F3F2F1"
        GRIS_MEDIO = "E1DFDD"
        
        @staticmethod
        def fill_encabezado_principal():
            return PatternFill(start_color=EstilosModernos.AZUL_PRINCIPAL, 
                              end_color=EstilosModernos.AZUL_PRINCIPAL, 
                              fill_type="solid")
        
        @staticmethod
        def fuente_encabezado():
            return Font(bold=True, size=11, color="FFFFFF", name='Arial')
        
        @staticmethod
        def fill_verde_sutil():
            return PatternFill(start_color=EstilosModernos.VERDE_SUTIL, 
                              end_color=EstilosModernos.VERDE_SUTIL, 
                              fill_type="solid")
        
        @staticmethod
        def fill_amarillo_sutil():
            return PatternFill(start_color=EstilosModernos.AMARILLO_SUTIL, 
                              end_color=EstilosModernos.AMARILLO_SUTIL, 
                              fill_type="solid")
        
        @staticmethod
        def fill_rojo_sutil():
            return PatternFill(start_color=EstilosModernos.ROJO_SUTIL, 
                              end_color=EstilosModernos.ROJO_SUTIL, 
                              fill_type="solid")
        
        @staticmethod
        def borde_sutil():
            return Border(
                left=Side(style='thin', color="E1DFDD"),
                right=Side(style='thin', color="E1DFDD"),
                top=Side(style='thin', color="E1DFDD"),
                bottom=Side(style='thin', color="E1DFDD")
            )
        
        @staticmethod
        def borde_medio():
            return Border(
                left=Side(style='medium', color="0054A6"),
                right=Side(style='medium', color="0054A6"),
                top=Side(style='medium', color="0054A6"),
                bottom=Side(style='medium', color="0054A6")
            )
        
        @staticmethod
        def alineacion_centro():
            return Alignment(horizontal="center", vertical="center", wrap_text=True)

def agregar_hoja_permisos(statistics_path="temp/statistics.json",
                         salida_excel="outputs/Reporte_Asistencias_Completo.xlsx"):
    """
    Agrega una hoja de resumen de permisos al reporte.
    MODIFICACIÓN: Diseño mejorado para reducir fatiga visual, bordes definidos y tamaño de fuente reducido.
    """
    print("\n📘 Agregando hoja 'Permisos' (diseño mejorado)...")

    try:
        wb = load_workbook(salida_excel)
    except FileNotFoundError:
        print(f"❌ No se encontró el archivo {salida_excel}")
        return

    # Crear o limpiar hoja existente
    if "Permisos" in wb.sheetnames:
        ws = wb["Permisos"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Permisos")

    # Cargar datos de statistics
    try:
        with open(statistics_path, "r", encoding="utf-8") as f:
            statistics = json.load(f)
    except Exception as e:
        print(f"❌ Error al leer el archivo de statistics: {e}")
        return

    if not statistics:
        print("❌ No hay datos para generar la hoja de permisos")
        ws.append(["No hay permisos para mostrar en el período analizado"])
        wb.save(salida_excel)
        return

    # Convertir a DataFrame para procesamiento
    df = pd.DataFrame(statistics)
    
    # Normalizar columnas
    df.columns = [col.strip().title() for col in df.columns]
    
    # Filtrar solo registros con permisos
    df_permisos = df[df.get("Permiso", "") == "Sí"].copy()
    
    if df_permisos.empty:
        print("❌ No hay permisos para mostrar")
        ws.append(["No hay permisos para mostrar en el período analizado"])
        wb.save(salida_excel)
        return

    # Crear columna de fecha
    df_permisos["Fecha"] = pd.to_datetime(df_permisos[["Año", "Mes", "Dia"]].rename(columns={
        "Año": "year", "Mes": "month", "Dia": "day"
    }), errors="coerce")
    
    df_permisos = df_permisos.dropna(subset=["Fecha"])
    
    # Normalizar IDs
    df_permisos["Cedula"] = df_permisos.get("Number_Id", pd.Series([""]*len(df_permisos))).fillna("").astype(str)

    # Determinar rango de fechas
    if df_permisos.empty:
        print("❌ DataFrame vacío después de procesar")
        return
        
    fecha_inicio = df_permisos["Fecha"].min()
    fecha_fin = df_permisos["Fecha"].max()
    print(f"🗓️ Rango de permisos detectado: {fecha_inicio.date()} → {fecha_fin.date()}")

    # =========================================================================
    # ESTILOS MEJORADOS - REDUCCIÓN DE FATIGA VISUAL
    # =========================================================================
    
    # Colores para fondos
    header_fill = PatternFill(start_color="0054A6", end_color="0054A6", fill_type="solid")  # Azul principal
    subheader_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")  # Azul más claro
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Azul muy claro
    light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Gris claro
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanco
    
    # Fuentes
    header_font_large = Font(bold=True, size=14, color="FFFFFF", name='Arial')
    header_font_medium = Font(bold=True, size=11, color="FFFFFF", name='Arial')
    header_font_small = Font(bold=True, size=8, color="FFFFFF", name='Arial')  # 🆕 Tamaño 8 para días
    normal_font = Font(size=10, name='Arial')
    small_font = Font(size=9, name='Arial')  # 🆕 Para contenido de celdas
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"),
        right=Side(style='thin', color="BFBFBF"),
        top=Side(style='thin', color="BFBFBF"),
        bottom=Side(style='thin', color="BFBFBF")
    )
    
    medium_border = Border(
        left=Side(style='medium', color="0054A6"),
        right=Side(style='medium', color="0054A6"),
        top=Side(style='medium', color="0054A6"),
        bottom=Side(style='medium', color="0054A6")
    )
    
    thick_border = Border(
        left=Side(style='thick', color="0054A6"),
        right=Side(style='thick', color="0054A6"),
        top=Side(style='thick', color="0054A6"),
        bottom=Side(style='thick', color="0054A6")
    )
    
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Calcular semanas
    try:
        # Ajustar para que la primera semana comience en lunes
        inicio_semana = fecha_inicio - timedelta(days=fecha_inicio.weekday())
        
        # Ajustar para que la última semana termine en domingo
        fin_semana = fecha_fin + timedelta(days=(6 - fecha_fin.weekday()))
        
        # Generar todas las semanas completas en el rango
        semanas = []
        current_semana_inicio = inicio_semana
        while current_semana_inicio <= fin_semana:
            current_semana_fin = current_semana_inicio + timedelta(days=6)
            semanas.append((current_semana_inicio, current_semana_fin))
            current_semana_inicio = current_semana_fin + timedelta(days=1)

        # Filtrar semanas que tienen permisos
        semanas_con_permisos = []
        for inicio, fin in semanas:
            tiene_permisos = False
            for fecha in [inicio + timedelta(days=i) for i in range(7)]:
                if len(df_permisos[df_permisos["Fecha"].dt.date == fecha.date()]) > 0:
                    tiene_permisos = True
                    break
            
            if tiene_permisos:
                semanas_con_permisos.append((inicio, fin))

        print(f"📅 Semanas con permisos: {len(semanas_con_permisos)} de {len(semanas)} totales")

    except Exception as e:
        print(f"❌ Error calculando semanas: {e}")
        return

    # Obtener todos los tipos de permisos únicos
    tipos_permiso = df_permisos["Tipo_Permiso"].unique()
    tipos_permiso = [tipo for tipo in tipos_permiso if tipo and tipo != "N/A"]
    tipos_permiso.sort()
    
    print(f"📋 Tipos de permisos encontrados: {len(tipos_permiso)}")
    for tipo in tipos_permiso:
        print(f"   - {tipo}")

    fila_actual = 1

    # Para cada tipo de permiso, crear una tabla
    for tipo_permiso in tipos_permiso:
        print(f"📊 Procesando tabla para: {tipo_permiso}")
        
        # Filtrar permisos de este tipo
        df_tipo = df_permisos[df_permisos["Tipo_Permiso"] == tipo_permiso].copy()
        
        # Agrupar empleados con permisos de este tipo
        empleados_tipo = {}
        
        for _, permiso in df_tipo.iterrows():
            try:
                cedula = permiso["Cedula"]
                nombre = permiso.get("Nombre", "")
                cargo = permiso.get("Cargo", "DOCENTE")
                fecha_permiso = permiso["Fecha"].date()
                
                if cedula not in empleados_tipo:
                    empleados_tipo[cedula] = {
                        "nombre": nombre,
                        "cargo": cargo,
                        "dias_permiso": set()
                    }
                
                # Agregar día de permiso
                empleados_tipo[cedula]["dias_permiso"].add(fecha_permiso)
                
            except Exception as e:
                print(f"⚠️ Error procesando permiso: {e}")
                continue

        if not empleados_tipo:
            continue

        # =========================================================================
        # ENCABEZADO DE LA TABLA - DISEÑO MEJORADO
        # =========================================================================
        
        # TÍTULO PRINCIPAL DE LA TABLA
        total_columnas = 4 + len(semanas_con_permisos) * 7  # 4 columnas fijas + semanas
        ws.merge_cells(f'A{fila_actual}:{get_column_letter(total_columnas)}{fila_actual}')
        cell = ws.cell(row=fila_actual, column=1, value=f"TIPO DE PERMISO: {tipo_permiso}")
        cell.fill = header_fill
        cell.font = header_font_large
        cell.alignment = center
        cell.border = thick_border
        fila_actual += 1

        # INFORMACIÓN DEL PERÍODO
        ws.merge_cells(f'A{fila_actual}:{get_column_letter(total_columnas)}{fila_actual}')
        cell = ws.cell(row=fila_actual, column=1, 
                      value=f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')} | Empleados: {len(empleados_tipo)} | Total de días con permiso: {sum(len(info['dias_permiso']) for info in empleados_tipo.values())}")
        cell.fill = subheader_fill
        cell.font = Font(bold=True, size=10, color="FFFFFF", name='Arial')
        cell.alignment = center
        cell.border = medium_border
        fila_actual += 1

        # ENCABEZADOS FIJOS (Documento, Nombre, Cargo, Total Días)
        headers_fijos = ["Documento", "Nombre", "Cargo", "Total Días"]
        start_col_fijos = 1
        
        for col_idx, header in enumerate(headers_fijos, start_col_fijos):
            cell = ws.cell(row=fila_actual, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font_medium
            cell.alignment = center
            cell.border = medium_border

        current_col = len(headers_fijos) + 1
        
        # ENCABEZADOS DE SEMANAS - CON FUENTE MÁS PEQUEÑA
        for i, (inicio, fin) in enumerate(semanas_con_permisos, 1):
            col_inicio = get_column_letter(current_col)
            col_fin = get_column_letter(current_col + 6)
            
            # Encabezado principal de semana
            ws.merge_cells(f"{col_inicio}{fila_actual}:{col_fin}{fila_actual}")
            cell = ws[f"{col_inicio}{fila_actual}"]
            cell.value = f"Semana {i}"
            cell.fill = subheader_fill
            cell.font = header_font_small  # 🆕 Tamaño 8
            cell.alignment = center
            cell.border = medium_border

            # Encabezados de días individuales
            dias_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb"]
            for j in range(6):
                fecha_dia = inicio + timedelta(days=j)
                dia_con_fecha = f"{dias_semana[j]}\n{fecha_dia.strftime('%d/%m')}"  # Formato más compacto
                cell = ws.cell(row=fila_actual + 1, column=current_col + j, value=dia_con_fecha)
                cell.fill = subheader_fill
                cell.font = header_font_small  # 🆕 Tamaño 8
                cell.alignment = center
                cell.border = medium_border

            # Columna "Total" de la semana
            cell = ws.cell(row=fila_actual + 1, column=current_col + 6, value="Total")
            cell.fill = subheader_fill
            cell.font = header_font_small  # 🆕 Tamaño 8
            cell.alignment = center
            cell.border = medium_border

            current_col += 7

        fila_actual += 2  # Saltar a la siguiente fila para los datos

        # =========================================================================
        # DATOS DE EMPLEADOS - DISEÑO MEJORADO
        # =========================================================================
        
        # Llenar datos de empleados para este tipo de permiso
        for idx, (cedula, info) in enumerate(empleados_tipo.items()):
            try:
                # Alternar colores de fila para mejor legibilidad
                if idx % 2 == 0:
                    row_fill = white_fill
                else:
                    row_fill = light_gray_fill
                
                # Escribir información básica del empleado
                ws.cell(row=fila_actual, column=1, value=cedula).border = thin_border
                ws.cell(row=fila_actual, column=1).fill = row_fill
                ws.cell(row=fila_actual, column=1).font = small_font
                
                ws.cell(row=fila_actual, column=2, value=info["nombre"]).border = thin_border
                ws.cell(row=fila_actual, column=2).fill = row_fill
                ws.cell(row=fila_actual, column=2).font = small_font
                ws.cell(row=fila_actual, column=2).alignment = left
                
                ws.cell(row=fila_actual, column=3, value=info["cargo"]).border = thin_border
                ws.cell(row=fila_actual, column=3).fill = row_fill
                ws.cell(row=fila_actual, column=3).font = small_font
                ws.cell(row=fila_actual, column=3).alignment = left

                current_col_data = 4
                total_permisos_empleado = len(info["dias_permiso"])
                
                # Celda de total de días de permiso
                total_dias_cell = ws.cell(row=fila_actual, column=4, value=total_permisos_empleado)
                total_dias_cell.alignment = center
                total_dias_cell.border = thin_border
                total_dias_cell.font = Font(bold=True, size=9)
                total_dias_cell.fill = light_blue_fill  # Destacar total

                current_col_data = 5  # Empezar en la columna 5 (después de "Total Días")
                
                # Procesar cada semana
                for inicio, fin in semanas_con_permisos:
                    permisos_semana = 0
                    
                    for j in range(6):  # Lunes a Sábado
                        fecha_dia = inicio + timedelta(days=j)
                        
                        # Verificar si este día tiene permiso de este tipo
                        if fecha_dia.date() in info["dias_permiso"]:
                            valor = "X"
                            permisos_semana += 1
                            cell_fill = light_blue_fill  # Azul claro para días con permiso
                        else:
                            valor = ""
                            cell_fill = row_fill
                        
                        # Escribir celda
                        cell = ws.cell(row=fila_actual, column=current_col_data + j, value=valor)
                        cell.alignment = center
                        cell.border = thin_border
                        cell.font = small_font
                        cell.fill = cell_fill

                    # Celda de total de la semana
                    total_cell = ws.cell(row=fila_actual, column=current_col_data + 6, value=permisos_semana)
                    total_cell.alignment = center
                    total_cell.border = thin_border
                    total_cell.font = Font(bold=True, size=9)
                    if permisos_semana > 0:
                        total_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Amarillo claro
                    else:
                        total_cell.fill = row_fill

                    current_col_data += 7

                fila_actual += 1

            except Exception as e:
                print(f"⚠️ Error procesando empleado {cedula}: {e}")
                continue

        # =========================================================================
        # BORDE FINAL DE LA TABLA - DEFINIR LÍMITES VISUALES
        # =========================================================================
        
        # Aplicar borde grueso alrededor de toda la tabla
        primera_fila_tabla = fila_actual - len(empleados_tipo) - 2  # -2 por los encabezados
        ultima_fila_tabla = fila_actual - 1
        
        for row in range(primera_fila_tabla, ultima_fila_tabla + 1):
            for col in range(1, total_columnas + 1):
                cell = ws.cell(row=row, column=col)
                # Si no tiene borde, aplicar borde delgado
                if not cell.border.left.style:
                    cell.border = thin_border
        
        # Aplicar borde exterior grueso a toda la tabla
        for col in range(1, total_columnas + 1):
            # Borde superior
            ws.cell(row=primera_fila_tabla, column=col).border = Border(
                top=Side(style='thick', color="0054A6"),
                left=ws.cell(row=primera_fila_tabla, column=col).border.left,
                right=ws.cell(row=primera_fila_tabla, column=col).border.right,
                bottom=ws.cell(row=primera_fila_tabla, column=col).border.bottom
            )
            # Borde inferior
            ws.cell(row=ultima_fila_tabla, column=col).border = Border(
                bottom=Side(style='thick', color="0054A6"),
                left=ws.cell(row=ultima_fila_tabla, column=col).border.left,
                right=ws.cell(row=ultima_fila_tabla, column=col).border.right,
                top=ws.cell(row=ultima_fila_tabla, column=col).border.top
            )
        
        for row in range(primera_fila_tabla, ultima_fila_tabla + 1):
            # Borde izquierdo
            ws.cell(row=row, column=1).border = Border(
                left=Side(style='thick', color="0054A6"),
                top=ws.cell(row=row, column=1).border.top,
                right=ws.cell(row=row, column=1).border.right,
                bottom=ws.cell(row=row, column=1).border.bottom
            )
            # Borde derecho
            ws.cell(row=row, column=total_columnas).border = Border(
                right=Side(style='thick', color="0054A6"),
                top=ws.cell(row=row, column=total_columnas).border.top,
                left=ws.cell(row=row, column=total_columnas).border.left,
                bottom=ws.cell(row=row, column=total_columnas).border.bottom
            )

        # Agregar espacio entre tablas de diferentes tipos de permisos
        fila_actual += 3

    # =========================================================================
    # AJUSTES FINALES DE FORMATO
    # =========================================================================
    
    # Ajustar anchos de columna
    try:
        ws.column_dimensions["A"].width = 12  # Documento
        ws.column_dimensions["B"].width = 25  # Nombre
        ws.column_dimensions["C"].width = 20  # Cargo
        ws.column_dimensions["D"].width = 8   # Total Días
        
        # Ajustar anchos para las columnas de semanas (más estrechas)
        max_col = 4 + len(semanas_con_permisos) * 7
        for col in range(5, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 6  # 🆕 Columnas más estrechas

        # Congelar paneles para las primeras filas (encabezados)
        if fila_actual > 3:
            ws.freeze_panes = "A3"

    except Exception as e:
        print(f"⚠️ Error ajustando formato: {e}")

    # Guardar
    wb.save(salida_excel)
    print(f"✅ Hoja 'Permisos' creada exitosamente en: {salida_excel}")
    print(f"📊 Total tipos de permisos procesados: {len(tipos_permiso)}")
    print(f"📅 Semanas mostradas: {len(semanas_con_permisos)}")
    print("🎨 DISEÑO MEJORADO: Tamaño de fuente 8 en días, bordes definidos, colores alternados")
    print("👁️  REDUCCIÓN DE FATIGA VISUAL: Filas alternadas, bordes claros, contraste adecuado")