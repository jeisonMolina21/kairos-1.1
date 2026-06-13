# reports/excel_resumen.py (MODIFICADO - MUESTRA TIPO DE PERMISO ESPECÍFICO)
import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import requests

try:
    from reports.styles import EstilosInstitucionales
except ImportError:
    pass

def obtener_festivos(año):
    """Obtiene los días festivos para Colombia del año específico"""
    try:
        url = f"https://date.nager.at/api/v3/PublicHolidays/{año}/CO"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        festivos = response.json()
        
        festivos_dates = []
        for festivo in festivos:
            fecha = datetime.strptime(festivo['date'], '%Y-%m-%d').date()
            if fecha.weekday() < 6:
                festivos_dates.append(fecha)
        
        return festivos_dates
    except Exception as e:
        print(f"⚠️ No se pudieron obtener los festivos: {e}")
        return []

def normalizar_cargo(cargo):
    """
    Normaliza los cargos para unificar variaciones
    """
    if not cargo or cargo == "N/A":
        return "DOCENTE"
    
    cargo = cargo.upper().strip()
    
    # Diccionario de normalización
    normalizacion = {
        'COORDINADORA': 'COORDINADOR',
        'COORDINADOR': 'COORDINADOR',
        'DIRECTORA': 'DIRECTOR', 
        'DIRECTOR': 'DIRECTOR',
        'ASISTENTE': 'ASISTENTE',
        'LIDER': 'LIDER',
        'ASESOR COMERCIAL': 'ASESOR COMERCIAL',
        'AUXILIAR ADMINISTRATIVO Y MENSAJERIA': 'AUXILIAR ADMINISTRATIVO',
        'RECTORA': 'RECTOR',
        'AUXILIAR MANTENIMIENTO, SERVICIOS GENERALES Y MENSAJERIA': 'AUXILIAR MANTENIMIENTO',
        'RECEPCIONISTA': 'RECEPCIONISTA',
        'DISEÑADOR GRAFICO': 'DISEÑADOR GRÁFICO',
        'VICERRECTORA': 'VICERRECTOR',
        'AUXILIAR': 'AUXILIAR',
        'COORDINADOR DE COMUNICACIONES': 'COORDINADOR',
        'ANALISTA': 'ANALISTA',
        'PRACTICANTE UNIVERSITARIA (PSICOLOGIA)': 'PRACTICANTE',
        'ANALISTA DE EGRESADOS': 'ANALISTA',
        'COORDINADORA ACADEMICA': 'COORDINADOR',
        'AUXILIAR SERVICIOS GENERALES': 'AUXILIAR',
        'DISEÑADOR GRÁFICO E-LEARNING': 'DISEÑADOR GRÁFICO',
        'ADMINISTRADORA PLATAFORMA LMS': 'ADMINISTRADOR PLATAFORMA',
        'DISEÑADOR INSTRUCCIONAL': 'DISEÑADOR INSTRUCCIONAL',
        'ASISTENTE DE BIBLIOTECA': 'ASISTENTE',
        'DIRECTOR DE INVESTIGACIONES': 'DIRECTOR',
        'AUXILIAR DE ADMISIONES': 'AUXILIAR',
        'SECRETARIA GENERAL': 'SECRETARIO GENERAL'
    }
    
    # Buscar coincidencias exactas primero
    if cargo in normalizacion:
        return normalizacion[cargo]
    
    # Buscar coincidencias parciales
    for key, value in normalizacion.items():
        if key in cargo:
            return value
    
    # Si no encuentra coincidencia, devolver el cargo original
    return cargo

def generar_excel_resumen_semanal(resumen_path="temp/resumen_asistencias.json",
                                  statistics_path="temp/statistics.json",
                                  salida_path="outputs/Reporte_Asistencias_Completo.xlsx"):
    """
    Genera la hoja de resumen semanal del reporte de asistencias.
    MODIFICACIÓN: Ahora muestra el tipo específico de permiso (ej: "Permiso: LR - Incapacidad medica")
    """
    print("\n📘 Generando hoja: Resumen semanal...")

    try:
        with open(resumen_path, "r", encoding="utf-8") as f:
            resumen = json.load(f)
    except Exception as e:
        print(f"❌ Error al cargar el archivo de resumen: {e}")
        return

    # Cargar statistics para obtener cargos y contar permisos
    try:
        with open(statistics_path, "r", encoding="utf-8") as f:
            statistics = json.load(f)
    except Exception as e:
        print(f"❌ Error al cargar statistics: {e}")
        statistics = []

    if not resumen:
        print("❌ No hay datos para generar el resumen")
        return

    # Crear DataFrame de manera robusta
    try:
        df = pd.DataFrame(resumen)
        
        # Verificar columnas críticas
        if 'Fecha' not in df.columns:
            print("❌ No se encuentra la columna 'Fecha' en los datos")
            return
            
        if 'Empleado' not in df.columns:
            print("❌ No se encuentra la columna 'Empleado' en los datos")
            return
            
        # Normalizar nombres de columnas
        df.columns = [col.strip().title() for col in df.columns]
        
        # Convertir fecha de manera segura
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
        df = df.dropna(subset=["Fecha"])
        
        # Normalizar cédula
        if 'Cedula' in df.columns:
            df["Cedula"] = df["Cedula"].fillna("").astype(str)
        else:
            df["Cedula"] = ""
            
    except Exception as e:
        print(f"❌ Error procesando DataFrame: {e}")
        return

    if df.empty:
        print("❌ DataFrame vacío después de procesar")
        return
        
    # Obtener rango de fechas de manera segura
    try:
        fecha_inicio = df["Fecha"].min()
        fecha_fin = df["Fecha"].max()
        print(f"🗓️ Rango detectado en resumen: {fecha_inicio.date()} → {fecha_fin.date()}")
    except Exception as e:
        print(f"❌ Error obteniendo rango de fechas: {e}")
        return

    # Obtener festivos para los años en el rango
    festivos = []
    try:
        años = range(fecha_inicio.year, fecha_fin.year + 1)
        for año in años:
            festivos.extend(obtener_festivos(año))
    except Exception as e:
        print(f"⚠️ Error obteniendo festivos: {e}")

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Semanal"

    # Estilos Institucionales
    header_fill = EstilosInstitucionales.fill_azul_principal()
    header_font = EstilosInstitucionales.fuente_encabezado_columna()
    center = EstilosInstitucionales.alineacion_centro()
    border = EstilosInstitucionales.borde_inferior_blanco()
    thin_border = EstilosInstitucionales.borde_horizontal_sutil()
    
    col_header_fill = EstilosInstitucionales.fill_azul_medio()
    normal_font = EstilosInstitucionales.fuente_normal()
    
    fill_total_oscuro = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
    font_total = Font(bold=True, size=10, color="FFFFFF", name=EstilosInstitucionales.FONT_NAME)
    
    green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    blue_fill = PatternFill(start_color="D6E4FF", end_color="D6E4FF", fill_type="solid")
    festivo_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    # Crear mapeo de cargos desde statistics
    mapeo_cargos = {}
    mapeo_permisos = {}
    
    # NUEVO: Crear mapeo de tipos de permisos por empleado y fecha
    mapeo_tipos_permisos = {}
    
    for stat in statistics:
        try:
            number_id = str(stat.get("number_id", "")).strip()
            cargo = stat.get("cargo", "DOCENTE").strip()
            permiso = stat.get("permiso", "No")
            tipo_permiso = stat.get("tipo_permiso", "N/A")
            año = stat.get("año")
            mes = stat.get("mes")
            dia = stat.get("dia")
            
            if number_id:
                cargo_normalizado = normalizar_cargo(cargo)
                mapeo_cargos[number_id] = cargo_normalizado
                
                if permiso == "Sí":
                    if number_id not in mapeo_permisos:
                        mapeo_permisos[number_id] = 0
                    mapeo_permisos[number_id] += 1
                    
                    # NUEVO: Guardar tipo de permiso por fecha
                    if año and mes and dia:
                        try:
                            fecha_str = f"{año}-{str(mes).zfill(2)}-{str(dia).zfill(2)}"
                            mapeo_tipos_permisos[(number_id, fecha_str)] = tipo_permiso
                        except Exception:
                            continue
        except Exception:
            continue

    print(f"📊 Cargos normalizados: {len(mapeo_cargos)} empleados")
    print(f"📋 Tipos de permisos mapeados: {len(mapeo_tipos_permisos)} registros")

    # Headers fijos
    headers_fijos = ["Documento", "Nombre", "Cargo"]
    for col_idx, header in enumerate(headers_fijos, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    current_col = len(headers_fijos) + 1
    
    # Calcular semanas de manera robusta
    try:
        # Ajustar para que la primera semana comience en lunes
        inicio_semana = fecha_inicio - timedelta(days=fecha_inicio.weekday())
        
        # Ajustar para que la última semana termine en domingo
        fin_semana = fecha_fin + timedelta(days=(6 - fecha_fin.weekday()))
        
        # Generar todas las semanas completas en el rango
        semanas = []
        current_semana_inicio = inicio_semana
        while current_semana_inicio <= fin_semana:
            current_semana_fin = current_semana_inicio + timedelta(days=6)
            semanas.append((current_semana_inicio, current_semana_fin))
            current_semana_inicio = current_semana_fin + timedelta(days=1)

        # Filtrar semanas que tienen registros en el resumen
        semanas_con_registros = []
        for inicio, fin in semanas:
            tiene_registros = False
            for fecha in [inicio + timedelta(days=i) for i in range(7)]:
                if len(df[df["Fecha"].dt.date == fecha.date()]) > 0:
                    tiene_registros = True
                    break
            
            if tiene_registros:
                semanas_con_registros.append((inicio, fin))

        print(f"📅 Semanas con registros en resumen: {len(semanas_con_registros)} de {len(semanas)} totales")

    except Exception as e:
        print(f"❌ Error calculando semanas: {e}")
        return

    # Escribir encabezados de semanas
    for i, (inicio, fin) in enumerate(semanas_con_registros, 1):
        col_inicio = get_column_letter(current_col)
        col_fin = get_column_letter(current_col + 6)
        
        ws.merge_cells(f"{col_inicio}1:{col_fin}1")
        cell = ws[f"{col_inicio}1"]
        cell.value = f"Semana {i}\n{inicio.strftime('%d/%m')} - {fin.strftime('%d/%m')}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

        # Encabezados de días con fecha completa
        dias_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb"]
        for j in range(6):
            fecha_dia = inicio + timedelta(days=j)
            dia_con_fecha = f"{dias_semana[j]} {fecha_dia.strftime('%d/%m')}"
            cell = ws.cell(row=2, column=current_col + j, value=dia_con_fecha)
            cell.fill = col_header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # Columna "Total" de la semana
        cell = ws.cell(row=2, column=current_col + 6, value="Total")
        cell.fill = fill_total_oscuro
        cell.font = font_total
        cell.alignment = center
        cell.border = border

        current_col += 7

    # Agregar columna "Total Permisos en el Mes"
    col_permisos_inicio = current_col
    ws.merge_cells(start_row=1, start_column=col_permisos_inicio, end_row=1, end_column=col_permisos_inicio)
    cell = ws.cell(row=1, column=col_permisos_inicio, value="Total Permisos en el Mes")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

    cell = ws.cell(row=2, column=col_permisos_inicio, value="Permisos")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

    # Agregar columna "Total Recargo Nocturno"
    col_recargo_inicio = current_col + 1
    ws.merge_cells(start_row=1, start_column=col_recargo_inicio, end_row=1, end_column=col_recargo_inicio)
    cell = ws.cell(row=1, column=col_recargo_inicio, value="Total Recargo Nocturno")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

    cell = ws.cell(row=2, column=col_recargo_inicio, value="Horas Nocturnas")
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

    fila_actual = 3
    
    # Agrupar empleados de manera robusta
    try:
        empleados_info = {}
        for empleado in df["Empleado"].unique():
            data_emp = df[df["Empleado"] == empleado]
            cedula = data_emp["Cedula"].iloc[0] if not data_emp.empty and "Cedula" in data_emp.columns else ""
            cargo = mapeo_cargos.get(cedula, "DOCENTE")
            total_permisos = mapeo_permisos.get(cedula, 0)
            
            empleados_info[empleado] = {
                "cedula": cedula,
                "cargo": cargo,
                "total_permisos": total_permisos,
                "data": data_emp
            }

        # Ordenar empleados: docentes al final
        def ordenar_empleado(item):
            empleado, info = item
            cargo = info["cargo"]
            es_docente = "DOCENTE" in cargo.upper()
            return (1 if es_docente else 0, empleado.upper())

        empleados_ordenados = sorted(empleados_info.items(), key=ordenar_empleado)

    except Exception as e:
        print(f"❌ Error agrupando empleados: {e}")
        return

    # Función robusta para obtener horas trabajadas
    def obtener_horas_trabajadas(registro):
        """Obtiene las horas trabajadas de un registro de manera robusta"""
        try:
            # Buscar en diferentes nombres de columnas
            for col_name in ['Horas_Trabajadas', 'Horas Trabajadas', 'horas_trabajadas']:
                if col_name in registro:
                    horas = registro[col_name]
                    if pd.isna(horas) or horas == "" or horas is None:
                        return 0.0
                    
                    if isinstance(horas, (int, float)):
                        return float(horas)
                    
                    if isinstance(horas, str):
                        if ":" in horas:
                            try:
                                partes = horas.split(":")
                                if len(partes) == 3:
                                    h, m, s = map(int, partes)
                                    return h + m/60 + s/3600
                                elif len(partes) == 2:
                                    h, m = map(int, partes)
                                    return h + m/60
                            except (ValueError, TypeError):
                                return 0.0
                        try:
                            return float(horas.replace(",", "."))
                        except (ValueError, TypeError):
                            return 0.0
            return 0.0
        except Exception:
            return 0.0

    # Llenar datos de empleados
    for empleado, info in empleados_ordenados:
        try:
            data_emp = info["data"]
            cedula = info["cedula"]
            cargo = info["cargo"]
            total_permisos = info["total_permisos"]
            
            # Escribir información básica del empleado
            ws.row_dimensions[fila_actual].height = 18
            for c in range(1, 4):
                cell = ws.cell(row=fila_actual, column=c)
                cell.border = thin_border
                cell.font = normal_font
                cell.alignment = center
                if (fila_actual % 2 == 1):
                    cell.fill = EstilosInstitucionales.fill_fila_par()

            ws.cell(row=fila_actual, column=1, value=cedula)
            ws.cell(row=fila_actual, column=2, value=empleado)
            ws.cell(row=fila_actual, column=3, value=cargo)

            current_col = 4
            recargo_nocturno_total_empleado = 0.0
            
            # O(1) Lookup: Convertir data_emp a un diccionario indexado por fecha
            # Extraer columnas relevantes de data_emp. Asumimos la primera fila por día.
            # Convertimos la columna Fecha a date para la clave
            registros_por_dia = {}
            for _, row in data_emp.iterrows():
                try:
                    fecha_val = row["Fecha"].date() if pd.notna(row["Fecha"]) else None
                    if fecha_val and fecha_val not in registros_por_dia:
                        registros_por_dia[fecha_val] = row.to_dict()
                except Exception:
                    pass
            
            # Procesar cada semana
            for inicio, fin in semanas_con_registros:
                horas_semana = 0
                
                for j in range(6):  # Lunes a Sábado
                    fecha_dia = inicio + timedelta(days=j)
                    fecha_dia_date = fecha_dia.date()
                    
                    es_festivo = fecha_dia_date in festivos
                    
                    # Filtrado O(1) del día
                    registro = registros_por_dia.get(fecha_dia_date)
                    
                    valor = ""
                    fill_color = None
                    
                    if es_festivo:
                        # Festivos suman 8 horas completas
                        horas_festivo = 8.0
                        valor = f"{int(horas_festivo)}:{int((horas_festivo%1)*60):02d}:00 / Festivo"
                        fill_color = festivo_fill
                        horas_semana += horas_festivo
                    elif registro:
                        # registro ya es un dict de los datos de la fila
                        horas = obtener_horas_trabajadas(registro)
                        
                        # Obtener tipo de evento
                        tipo_evento = ""
                        for col_name in ['Tipo_Evento', 'Tipo Evento', 'tipo_evento']:
                            if col_name in registro:
                                tipo_evento = registro[col_name]
                                break
                        
                        horas_semana += horas
                        
                        # CALCULAR RECARGO NOCTURNO
                        from core.config import work_schedule
                        umbral_nocturno = work_schedule.HORA_INICIO_RECARGO_NOCTURNO
                        
                        for col_name in ['Hora_Salida', 'Hora Salida', 'hora_salida']:
                            if col_name in registro:
                                hora_salida_str = registro[col_name]
                                if isinstance(hora_salida_str, str) and ":" in hora_salida_str:
                                    try:
                                        partes = hora_salida_str.split(":")
                                        h_salida = int(partes[0])
                                        m_salida = int(partes[1]) if len(partes) > 1 else 0
                                        salida_decimal = h_salida + (m_salida / 60.0)
                                        if salida_decimal > umbral_nocturno:
                                            recargo_nocturno_total_empleado += (salida_decimal - umbral_nocturno)
                                    except:
                                        pass
                                break
                        
                        # MODIFICACIÓN PRINCIPAL: Mostrar tipo específico de permiso
                        if tipo_evento in ["Permiso", "Campaña", "Salida Institucional"] and horas > 0:
                            if tipo_evento == "Permiso":
                                # Buscar el tipo específico de permiso
                                fecha_str = fecha_dia.strftime("%Y-%m-%d")
                                tipo_permiso_especifico = mapeo_tipos_permisos.get((cedula, fecha_str), "Permiso")
                                if tipo_permiso_especifico == "N/A":
                                    tipo_permiso_especifico = "Permiso"
                                valor = f"{int(horas)}:{int((horas%1)*60):02d}:00 / {tipo_evento}: {tipo_permiso_especifico}"
                            else:
                                valor = f"{int(horas)}:{int((horas%1)*60):02d}:00 / {tipo_evento}"
                            fill_color = blue_fill
                        elif tipo_evento in ["Permiso", "Campaña", "Salida Institucional"]:
                            if tipo_evento == "Permiso":
                                # Buscar el tipo específico de permiso
                                fecha_str = fecha_dia.strftime("%Y-%m-%d")
                                tipo_permiso_especifico = mapeo_tipos_permisos.get((cedula, fecha_str), "Permiso")
                                if tipo_permiso_especifico == "N/A":
                                    tipo_permiso_especifico = "Permiso"
                                valor = f"{tipo_evento}: {tipo_permiso_especifico}"
                            else:
                                valor = tipo_evento
                            fill_color = blue_fill
                        elif horas > 0:
                            valor = f"{int(horas)}:{int((horas%1)*60):02d}:00"
                            if horas >= 8:
                                fill_color = green_fill
                            elif horas >= 6:
                                fill_color = yellow_fill
                            else:
                                fill_color = red_fill
                        else:
                            valor = "Sin registros"
                            fill_color = red_fill
                    else:
                        valor = "Sin registros"
                        fill_color = red_fill
                    
                    # Escribir celda
                    cell = ws.cell(row=fila_actual, column=current_col + j, value=valor)
                    cell.alignment = center
                    cell.border = thin_border
                    cell.font = normal_font
                    
                    if (fila_actual % 2 == 1):
                        cell.fill = EstilosInstitucionales.fill_fila_par()
                        
                    # Validar si tiene badge específico por texto (Permisos, etc)
                    bg, fg = EstilosInstitucionales.get_badge_por_texto(str(valor))
                    if bg:
                        cell.fill = bg
                        cell.font = fg
                    elif fill_color: # Retención de lógica para verde, amarillo, etc
                        cell.fill = fill_color

                # Celda de total de la semana
                total_cell = ws.cell(row=fila_actual, column=current_col + 6, 
                                   value=f"{int(horas_semana)}:{int((horas_semana%1)*60):02d}:00")
                total_cell.alignment = center
                total_cell.border = thin_border
                total_cell.font = Font(bold=True, name=EstilosInstitucionales.FONT_NAME)
                
                if (fila_actual % 2 == 1):
                    total_cell.fill = EstilosInstitucionales.fill_fila_par()

                if horas_semana >= 40:
                    total_cell.fill = green_fill
                elif horas_semana >= 30:
                    total_cell.fill = yellow_fill
                else:
                    total_cell.fill = red_fill

                current_col += 7

            # Celda de total de permisos
            permisos_cell = ws.cell(row=fila_actual, column=col_permisos_inicio, value=total_permisos)
            permisos_cell.alignment = center
            permisos_cell.border = thin_border
            permisos_cell.font = normal_font
            if (fila_actual % 2 == 1):
                permisos_cell.fill = EstilosInstitucionales.fill_fila_par()

            # Celda de total recargo nocturno
            horas_recargo = int(recargo_nocturno_total_empleado)
            minutos_recargo = int((recargo_nocturno_total_empleado % 1) * 60)
            recargo_valor = f"{horas_recargo}:{minutos_recargo:02d}:00" if recargo_nocturno_total_empleado > 0 else "0:00:00"
            recargo_cell = ws.cell(row=fila_actual, column=col_recargo_inicio, value=recargo_valor)
            recargo_cell.alignment = center
            recargo_cell.border = thin_border
            recargo_cell.font = normal_font
            if (fila_actual % 2 == 1):
                recargo_cell.fill = EstilosInstitucionales.fill_fila_par()

            fila_actual += 1

        except Exception as e:
            print(f"⚠️ Error procesando empleado {empleado}: {e}")
            continue

    # Ajustar anchos de columna
    try:
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 25
        for col in range(4, current_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        ws.column_dimensions[get_column_letter(col_permisos_inicio)].width = 12
        ws.column_dimensions[get_column_letter(col_recargo_inicio)].width = 15

        # Agregar filtros automáticos
        ws.auto_filter.ref = f"A1:{get_column_letter(3)}{fila_actual-1}"

        # Congelar paneles: Columnas 1-3 y filas 1-2. Intersección es D3.
        ws.freeze_panes = "D3"

    except Exception as e:
        print(f"⚠️ Error ajustando formato: {e}")

    # Guardar
    try:
        wb.save(salida_path)
        print(f"✅ Hoja 'Resumen Semanal' creada exitosamente en: {salida_path}")
        print(f"📊 Total empleados procesados: {len(empleados_ordenados)}")
        print(f"📅 Semanas mostradas: {len(semanas_con_registros)}")
        print("🎯 Festivos detectados y marcados en amarillo")
        print("🔵 Permisos, Campañas y Salidas Institucionales en azul")
        print("📝 NUEVO: Se muestran tipos específicos de permisos (ej: 'Permiso: LR - Incapacidad médica')")
        
    except Exception as e:
        print(f"❌ Error guardando archivo: {e}")