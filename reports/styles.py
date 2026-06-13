from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class EstilosModernos:
    AZUL_PRINCIPAL = "0054A6"
    AZUL_SECUNDARIO = "0078D4"
    VERDE_EXITO = "107C10"
    NARANJA_ALERTA = "D83B01"
    ROJO_CRITICO = "E81123"
    GRIS_CLARO = "F3F2F1"
    GRIS_MEDIO = "E1DFDD"
    BLANCO = "FFFFFF"
    NEGRO = "323130"
    EMERGENCIA = "E9D502"
    
    VERDE_SUTIL = "C6E0B4"
    AMARILLO_SUTIL = "FFE699"
    ROJO_SUTIL = "F8CBAD"
    
    @staticmethod
    def fuente_titulo():
        return Font(bold=True, size=16, color=EstilosModernos.BLANCO, name='Arial')
    
    @staticmethod
    def fuente_subtitulo():
        return Font(bold=True, size=12, color=EstilosModernos.BLANCO, name='Arial')
    
    @staticmethod
    def fuente_encabezado():
        return Font(bold=True, size=11, color=EstilosModernos.BLANCO, name='Arial')
    
    @staticmethod
    def fuente_normal():
        return Font(size=10, name='Arial')
    
    @staticmethod
    def fuente_destacada():
        return Font(bold=True, size=10, name='Arial')
    
    @staticmethod
    def alineacion_centro():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)
    
    @staticmethod
    def alineacion_izquierda():
        return Alignment(horizontal="left", vertical="center", wrap_text=False)
    
    @staticmethod
    def alineacion_derecha():
        return Alignment(horizontal="right", vertical="center", wrap_text=False)
    
    @staticmethod
    def borde_sutil():
        return Border(
            left=Side(style='thin', color=EstilosModernos.GRIS_MEDIO),
            right=Side(style='thin', color=EstilosModernos.GRIS_MEDIO),
            top=Side(style='thin', color=EstilosModernos.GRIS_MEDIO),
            bottom=Side(style='thin', color=EstilosModernos.GRIS_MEDIO)
        )
    
    @staticmethod
    def borde_destacado():
        return Border(
            left=Side(style='medium', color=EstilosModernos.AZUL_PRINCIPAL),
            right=Side(style='medium', color=EstilosModernos.AZUL_PRINCIPAL),
            top=Side(style='medium', color=EstilosModernos.AZUL_PRINCIPAL),
            bottom=Side(style='medium', color=EstilosModernos.AZUL_PRINCIPAL)
        )
    
    @staticmethod
    def fill_encabezado_principal():
        return PatternFill(start_color=EstilosModernos.AZUL_PRINCIPAL, 
                          end_color=EstilosModernos.AZUL_PRINCIPAL, 
                          fill_type="solid")
    
    @staticmethod
    def fill_encabezado_secundario():
        return PatternFill(start_color=EstilosModernos.AZUL_SECUNDARIO, 
                          end_color=EstilosModernos.AZUL_SECUNDARIO, 
                          fill_type="solid")
    
    @staticmethod
    def fill_fila_par():
        return PatternFill(start_color=EstilosModernos.GRIS_CLARO, 
                          end_color=EstilosModernos.GRIS_CLARO, 
                          fill_type="solid")
    
    @staticmethod
    def fill_verde():
        return PatternFill(start_color=EstilosModernos.VERDE_EXITO, 
                          end_color=EstilosModernos.VERDE_EXITO, 
                          fill_type="solid")
    
    @staticmethod
    def fill_naranja():
        return PatternFill(start_color=EstilosModernos.NARANJA_ALERTA, 
                          end_color=EstilosModernos.NARANJA_ALERTA, 
                          fill_type="solid")
    
    @staticmethod
    def fill_rojo():
        return PatternFill(start_color=EstilosModernos.ROJO_CRITICO, 
                          end_color=EstilosModernos.ROJO_CRITICO, 
                          fill_type="solid")
    
    @staticmethod
    def fill_verde_sutil():
        return PatternFill(start_color=EstilosModernos.VERDE_SUTIL, 
                          end_color=EstilosModernos.VERDE_SUTIL, 
                          fill_type="solid")
    
    @staticmethod
    def fill_amarillo_sutil():
        return PatternFill(start_color=EstilosModernos.AMARILLO_SUTIL, 
                          end_color=EstilosModernos.AMARILLO_SUTIL, 
                          fill_type="solid")
    
    @staticmethod
    def fill_rojo_sutil():
        return PatternFill(start_color=EstilosModernos.ROJO_SUTIL, 
                          end_color=EstilosModernos.ROJO_SUTIL, 
                          fill_type="solid")
    
    @staticmethod
    def fill_emergencia():
        return PatternFill(start_color=EstilosModernos.EMERGENCIA, 
                          end_color=EstilosModernos.EMERGENCIA, 
                          fill_type="solid")

class EstilosInstitucionales:
    """Nuevos estilos modernos, flat y profesionales (Calibri)"""
    AZUL_PRINCIPAL = "1A5EA8"
    AZUL_OSCURO = "0C447C"
    AZUL_CLARO = "E6F1FB"
    AZUL_MEDIO = "2E6DC4"
    BLANCO_OPACO = "B5D4F4" # Blanco al 80%
    
    BLANCO = "FFFFFF"
    TEXTO_NEGRO = "1A1A1A"
    
    # Badges Colors
    BADGE_SIN_REGISTRO_BG = "F1F5F9"
    BADGE_SIN_REGISTRO_FG = "64748B"
    BADGE_PERMISO_BG = "DBEAFE"
    BADGE_PERMISO_FG = "1A5EA8"
    BADGE_TARDE_BG = "FEF3C7"
    BADGE_TARDE_FG = "92400E"
    BADGE_EXCELENTE_BG = "D1FAE5"
    BADGE_EXCELENTE_FG = "065F46"
    BADGE_CRITICO_BG = "FEE2E2"
    BADGE_CRITICO_FG = "991B1B"
    BADGE_FESTIVO_BG = "FEF9ED"
    BADGE_FESTIVO_FG = "92400E"
    
    FILA_PAR_BG = "F7F9FD"
    BORDE_SUTIL_COLOR = "D1DCF0"
    
    FONT_NAME = 'Calibri'
    
    @staticmethod
    def fuente_titulo_principal():
        return Font(bold=True, size=13, color=EstilosInstitucionales.BLANCO, name=EstilosInstitucionales.FONT_NAME)
    
    @staticmethod
    def fuente_periodo():
        return Font(size=11, color=EstilosInstitucionales.BLANCO_OPACO, name=EstilosInstitucionales.FONT_NAME)
        
    @staticmethod
    def fuente_metadatos():
        return Font(size=10, color=EstilosInstitucionales.BLANCO_OPACO, name=EstilosInstitucionales.FONT_NAME)
    
    @staticmethod
    def fuente_encabezado_columna():
        return Font(bold=True, size=10, color=EstilosInstitucionales.BLANCO, name=EstilosInstitucionales.FONT_NAME)
    
    @staticmethod
    def fuente_normal():
        return Font(size=10, color=EstilosInstitucionales.TEXTO_NEGRO, name=EstilosInstitucionales.FONT_NAME)
    
    @staticmethod
    def fuente_bold_normal():
        return Font(bold=True, size=10, color=EstilosInstitucionales.TEXTO_NEGRO, name=EstilosInstitucionales.FONT_NAME)
    
    @staticmethod
    def alineacion_centro():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)
    
    @staticmethod
    def alineacion_izquierda():
        return Alignment(horizontal="left", vertical="center", wrap_text=False)
    
    @staticmethod
    def borde_inferior_blanco():
        return Border(bottom=Side(style='thin', color=EstilosInstitucionales.BLANCO))
        
    @staticmethod
    def borde_horizontal_sutil():
        return Border(
            bottom=Side(style='thin', color=EstilosInstitucionales.BORDE_SUTIL_COLOR),
            top=Side(style='thin', color=EstilosInstitucionales.BORDE_SUTIL_COLOR)
        )
        
    @staticmethod
    def fill_azul_principal():
        return PatternFill(start_color=EstilosInstitucionales.AZUL_PRINCIPAL, end_color=EstilosInstitucionales.AZUL_PRINCIPAL, fill_type="solid")
        
    @staticmethod
    def fill_azul_medio():
        return PatternFill(start_color=EstilosInstitucionales.AZUL_MEDIO, end_color=EstilosInstitucionales.AZUL_MEDIO, fill_type="solid")
        
    @staticmethod
    def fill_azul_oscuro():
        return PatternFill(start_color=EstilosInstitucionales.AZUL_OSCURO, end_color=EstilosInstitucionales.AZUL_OSCURO, fill_type="solid")
    
    @staticmethod
    def fill_fila_par():
        return PatternFill(start_color=EstilosInstitucionales.FILA_PAR_BG, end_color=EstilosInstitucionales.FILA_PAR_BG, fill_type="solid")

    @staticmethod
    def badge_sin_registro():
        return (PatternFill(start_color=EstilosInstitucionales.BADGE_SIN_REGISTRO_BG, end_color=EstilosInstitucionales.BADGE_SIN_REGISTRO_BG, fill_type="solid"),
                Font(bold=True, size=9, color=EstilosInstitucionales.BADGE_SIN_REGISTRO_FG, name=EstilosInstitucionales.FONT_NAME))

    @staticmethod
    def badge_permiso():
        return (PatternFill(start_color=EstilosInstitucionales.BADGE_PERMISO_BG, end_color=EstilosInstitucionales.BADGE_PERMISO_BG, fill_type="solid"),
                Font(bold=True, size=9, color=EstilosInstitucionales.BADGE_PERMISO_FG, name=EstilosInstitucionales.FONT_NAME))

    @staticmethod
    def badge_tarde():
        return (PatternFill(start_color=EstilosInstitucionales.BADGE_TARDE_BG, end_color=EstilosInstitucionales.BADGE_TARDE_BG, fill_type="solid"),
                Font(bold=True, size=9, color=EstilosInstitucionales.BADGE_TARDE_FG, name=EstilosInstitucionales.FONT_NAME))

    @staticmethod
    def badge_excelente():
        return (PatternFill(start_color=EstilosInstitucionales.BADGE_EXCELENTE_BG, end_color=EstilosInstitucionales.BADGE_EXCELENTE_BG, fill_type="solid"),
                Font(bold=True, size=9, color=EstilosInstitucionales.BADGE_EXCELENTE_FG, name=EstilosInstitucionales.FONT_NAME))

    @staticmethod
    def badge_critico():
        return (PatternFill(start_color=EstilosInstitucionales.BADGE_CRITICO_BG, end_color=EstilosInstitucionales.BADGE_CRITICO_BG, fill_type="solid"),
                Font(bold=True, size=9, color=EstilosInstitucionales.BADGE_CRITICO_FG, name=EstilosInstitucionales.FONT_NAME))

    @staticmethod
    def badge_festivo():
        return (PatternFill(start_color=EstilosInstitucionales.BADGE_FESTIVO_BG, end_color=EstilosInstitucionales.BADGE_FESTIVO_BG, fill_type="solid"),
                Font(bold=True, size=9, color=EstilosInstitucionales.BADGE_FESTIVO_FG, name=EstilosInstitucionales.FONT_NAME))
                
    @staticmethod
    def get_badge_por_texto(texto):
        texto = str(texto).lower()
        if "sin registro" in texto:
            return EstilosInstitucionales.badge_sin_registro()
        elif "permiso" in texto or "campaña" in texto or "salida institucional" in texto:
            return EstilosInstitucionales.badge_permiso()
        elif "tarde" in texto or "regular" in texto:
            return EstilosInstitucionales.badge_tarde()
        elif "tiempo" in texto or "excelente" in texto or "novedad" in texto or "cumplió" in texto or "completas" in texto:
            return EstilosInstitucionales.badge_excelente()
        elif "festivo" in texto:
            return EstilosInstitucionales.badge_festivo()
        elif "crítico" in texto or "mejorar" in texto or "00:00:00" in texto:
            return EstilosInstitucionales.badge_critico()
        # Fallback a normal
        return None, None